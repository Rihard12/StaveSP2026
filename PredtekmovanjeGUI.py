import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import logging
import openpyxl
import xlwings as xw
from rapidfuzz import process
from openpyxl.utils import column_index_from_string, get_column_letter
import numpy as np
from datetime import datetime

try:
    import pythoncom
    _HAS_PYTHONCOM = True
except ImportError:
    _HAS_PYTHONCOM = False

def get_base_dir():
    """Vrne mapo kjer je .exe (pri buildu) oz. kjer je gui.py (pri razvoju)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_dir()

# ──────────────────────────────────────────────
#  GLOBALNE SPREMENLJIVKE (nastavljive v GUI)
# ──────────────────────────────────────────────
NAPOVEDI_MAPA     = os.path.join(BASE_DIR, "TESTNapovediPredtekmovanje")
REZULTATI_FILE    = os.path.join(BASE_DIR, "TEST - SP 2026 - predtekmovanje - rezultati.xlsx")
LOG_FILE          = os.path.join(BASE_DIR, "log.txt")

IME_CELL          = "I3"
DOMACI_COL        = "I"
GOST_COL          = "K"
TOCKE_COL         = "L"
IME_COLUMN        = "B"

TEKME_FILE        = os.path.join(BASE_DIR, "SP 2026 - tekme.xlsx")
TEKME_SHEET_IDX   = 0    # Indeks lista za tekme
SKUPINE_SHEET_IDX = 1    # Indeks lista za skupine
IZLOCILNI_SHEET_IDX = 2  # Indeks lista za izločilne tekme (v TEKME_FILE)

# Izločilni del
NAPOVEDI_MAPA_IZL  = os.path.join(BASE_DIR, "NapovediIzlocilni")
REZULTATI_FILE_IZL = os.path.join(BASE_DIR, "SP 2026 - izlocilni - rezultati.xlsx")
IME_CELL_IZL       = "U3"
IME_COLUMN_IZL     = "B"

IMENA_VRSTICE     = {}
IMENA_VRSTICE_IZL = {}

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(message)s"
)


CONFIG_FILE = os.path.join(BASE_DIR, "config.json")

def nalozi_config():
    """Naloži nastavitve iz config.json, če obstaja."""
    global NAPOVEDI_MAPA, REZULTATI_FILE, TEKME_FILE, LOG_FILE
    global IME_CELL, DOMACI_COL, GOST_COL, TOCKE_COL, IME_COLUMN
    global TEKME_SHEET_IDX, SKUPINE_SHEET_IDX, IZLOCILNI_SHEET_IDX
    global NAPOVEDI_MAPA_IZL, REZULTATI_FILE_IZL, IME_CELL_IZL, IME_COLUMN_IZL
    if not os.path.exists(CONFIG_FILE):
        return
    try:
        import json
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        NAPOVEDI_MAPA       = cfg.get("NAPOVEDI_MAPA",       NAPOVEDI_MAPA)
        REZULTATI_FILE      = cfg.get("REZULTATI_FILE",      REZULTATI_FILE)
        TEKME_FILE          = cfg.get("TEKME_FILE",          TEKME_FILE)
        IME_CELL            = cfg.get("IME_CELL",            IME_CELL)
        DOMACI_COL          = cfg.get("DOMACI_COL",          DOMACI_COL)
        GOST_COL            = cfg.get("GOST_COL",            GOST_COL)
        TOCKE_COL           = cfg.get("TOCKE_COL",           TOCKE_COL)
        IME_COLUMN          = cfg.get("IME_COLUMN",          IME_COLUMN)
        TEKME_SHEET_IDX     = cfg.get("TEKME_SHEET_IDX",     TEKME_SHEET_IDX)
        SKUPINE_SHEET_IDX   = cfg.get("SKUPINE_SHEET_IDX",   SKUPINE_SHEET_IDX)
        IZLOCILNI_SHEET_IDX = cfg.get("IZLOCILNI_SHEET_IDX", IZLOCILNI_SHEET_IDX)
        NAPOVEDI_MAPA_IZL   = cfg.get("NAPOVEDI_MAPA_IZL",   NAPOVEDI_MAPA_IZL)
        REZULTATI_FILE_IZL  = cfg.get("REZULTATI_FILE_IZL",  REZULTATI_FILE_IZL)
        IME_CELL_IZL        = cfg.get("IME_CELL_IZL",        IME_CELL_IZL)
        IME_COLUMN_IZL      = cfg.get("IME_COLUMN_IZL",      IME_COLUMN_IZL)
    except Exception as e:
        print(f"Opozorilo: ni bilo mogoče naložiti config.json: {e}")

def shrani_config():
    """Shrani trenutne nastavitve v config.json."""
    import json
    cfg = {
        "NAPOVEDI_MAPA":       NAPOVEDI_MAPA,
        "REZULTATI_FILE":      REZULTATI_FILE,
        "TEKME_FILE":          TEKME_FILE,
        "IME_CELL":            IME_CELL,
        "DOMACI_COL":          DOMACI_COL,
        "GOST_COL":            GOST_COL,
        "TOCKE_COL":           TOCKE_COL,
        "IME_COLUMN":          IME_COLUMN,
        "TEKME_SHEET_IDX":     TEKME_SHEET_IDX,
        "SKUPINE_SHEET_IDX":   SKUPINE_SHEET_IDX,
        "IZLOCILNI_SHEET_IDX": IZLOCILNI_SHEET_IDX,
        "NAPOVEDI_MAPA_IZL":   NAPOVEDI_MAPA_IZL,
        "REZULTATI_FILE_IZL":  REZULTATI_FILE_IZL,
        "IME_CELL_IZL":        IME_CELL_IZL,
        "IME_COLUMN_IZL":      IME_COLUMN_IZL,
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

# Naloži config takoj ob zagonu
nalozi_config()

# ──────────────────────────────────────────────
#  POSLOVNA LOGIKA  (enako kot original)
# ──────────────────────────────────────────────

def vrni_tip(d, g):
    if d > g: return 1
    if d < g: return 2
    return 0

def napolni_slovar_imen():
    global IMENA_VRSTICE
    if os.path.isfile(REZULTATI_FILE):
        app = xw.App(visible=False)
        try:
            wb  = app.books.open(REZULTATI_FILE)
            ws  = wb.sheets[0]
            imena = ws.range(f"{IME_COLUMN}3").expand("down").value
            if imena:
                IMENA_VRSTICE = {ime.strip().title(): i for i, ime in enumerate(imena, start=3)}
            wb.close()
        finally:
            app.quit()

def obdelaj_tekmo(vrsticaNapovedi, goliDomaciRez, goliGostjeRez, stolpecRez, log_cb, progress_cb=None):
    tipRez = vrni_tip(goliDomaciRez, goliGostjeRez)
    files  = [f for f in os.listdir(NAPOVEDI_MAPA) if f.endswith(".xlsx") and not f.startswith("~")]
    skupaj = len(files)
    stat = {3: 0, 1: 0, 0: 0, "napake": 0}
    app = xw.App(visible=False)
    try:
        wb  = app.books.open(REZULTATI_FILE)
        ws  = wb.sheets[0]
        for i, file in enumerate(files):
            path = os.path.join(NAPOVEDI_MAPA, file)
            try:
                wbn = openpyxl.load_workbook(path, data_only=True)
                wsn = wbn.worksheets[0]
                ime  = wsn[IME_CELL].value.strip().title()
                vr   = IMENA_VRSTICE.get(ime)
                if vr is None: raise Exception("Ime ni najdeno!")
                gd = int(wsn[f"{DOMACI_COL}{vrsticaNapovedi}"].value)
                gg = int(wsn[f"{GOST_COL}{vrsticaNapovedi}"].value)
                tipN = vrni_tip(gd, gg)
                if gd == goliDomaciRez and gg == goliGostjeRez:
                    tocke = 3
                elif tipN == tipRez:
                    tocke = 1
                else:
                    tocke = 0
                ws[f"{stolpecRez}{vr}"].value = tocke
                wbn.close()
                stat[tocke] += 1
                logging.info(f"{ime} | {gd}:{gg} | {tocke} tock")
                log_cb(f"✓  {ime}: {tocke} točk")
            except Exception as e:
                stat["napake"] += 1
                logging.error(f"NAPAKA {file}: {e}")
                log_cb(f"✗  {file}: {e}")
            if progress_cb: progress_cb(int((i + 1) / skupaj * 100))
        wb.save(REZULTATI_FILE)
        wb.close()
    finally:
        app.quit()
    log_cb(f"\n── Povzetek ──────────────────────")
    log_cb(f"Obdelanih: {skupaj - stat['napake']}/{skupaj}  |  3 točke: {stat[3]}  |  1 točka: {stat[1]}  |  0 točk: {stat[0]}  |  Napake: {stat['napake']}")

def obdelaj_skupino(skupina, vrsticaNapoved, stolpecNapoved, stolpecRez, ekipe, log_cb, progress_cb=None):
    files = [f for f in os.listdir(NAPOVEDI_MAPA) if f.endswith(".xlsx") and not f.startswith("~")]
    ekipeSeznam = [x.strip().upper() for x in ekipe]
    skupaj = len(files)
    stat = {8: 0, 6: 0, 4: 0, 2: 0, 0: 0, "napake": 0}
    app = xw.App(visible=False)
    try:
        wb  = app.books.open(REZULTATI_FILE)
        ws  = wb.sheets[0]
        for i, file in enumerate(files):
            path = os.path.join(NAPOVEDI_MAPA, file)
            try:
                wbn = openpyxl.load_workbook(path, data_only=True)
                wsn = wbn.worksheets[0]
                ime = wsn[IME_CELL].value.strip().title()
                vr  = IMENA_VRSTICE.get(ime)
                if vr is None: raise Exception("Ime ni najdeno!")
                points = 0
                for idx in range(4):
                    napoved = wsn[f"{stolpecNapoved}{vrsticaNapoved + idx}"].value
                    napoved = str(napoved).strip()
                    match, procent = razreši_z_aliasi(napoved, ekipeSeznam, ekipe)
                    if procent < PRAG_VALIDACIJE:
                        log_cb(f"  ⚠  {ime}, mesto {idx+1}: '{napoved}' → '{match.title()}' ({procent:.0f}%) – preskočeno")
                        logging.warning(f"{ime} | Skupina {skupina} | '{napoved}' → '{match}' ({procent:.0f}%) pod pragom")
                        continue
                    if match == ekipeSeznam[idx]:
                        points += 2
                        col = get_column_letter(column_index_from_string(stolpecRez) + idx)
                        ws[f"{col}{vr}"].value = 2
                stat[points] = stat.get(points, 0) + 1
                logging.info(f"{ime} | Skupina {skupina} | {points} točk")
                log_cb(f"✓  {ime}: {points} točk")
                wbn.close()
            except Exception as e:
                stat["napake"] += 1
                logging.error(f"NAPAKA {file}: {e}")
                log_cb(f"✗  {file}: {e}")
            if progress_cb: progress_cb(int((i + 1) / skupaj * 100))
        wb.save(REZULTATI_FILE)
        wb.close()
    finally:
        app.quit()
    tocke_dist = ", ".join(f"{t}t: {n}" for t, n in sorted((k, v) for k, v in stat.items() if isinstance(k, int) and v > 0))
    log_cb(f"\n── Povzetek ──────────────────────")
    log_cb(f"Obdelanih: {skupaj - stat['napake']}/{skupaj}  |  {tocke_dist}  |  Napake: {stat['napake']}")

def napolni_imena(log_cb):
    vrstica = 3
    files = [f for f in os.listdir(NAPOVEDI_MAPA) if f.endswith(".xlsx") and not f.startswith("~")]
    app = xw.App(visible=False)
    try:
        wb  = app.books.open(REZULTATI_FILE)
        ws  = wb.sheets[0]
        for file in files:
            path = os.path.join(NAPOVEDI_MAPA, file)
            wbn  = None
            try:
                wbn = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ime = wbn.worksheets[0][IME_CELL].value.strip().title()
                wbn.close()
                ws[f"{IME_COLUMN}{vrstica}"].value = ime
                vrstica += 1
                log_cb(f"✓  {file}")
            except Exception as e:
                if wbn: wbn.close()
                logging.error(f"NAPAKA {file}: {e}")
                log_cb(f"✗  {file}: {e}")
        wb.save(REZULTATI_FILE)
        wb.close()
    finally:
        app.quit()

    log_cb(f"✓  Imena uspešno prebrana!")
    napolni_slovar_imen()

def preberi_aktivne_in_izpadle_ekipe():
    """
    Vrne (preostale_ekipe, vse_kdaj_omenjene) na podlagi SP_2026_-_tekme.xlsx (list Izlocilni).
    preostale_ekipe: množica imen ekip (UPPER), ki nastopajo v še neobdelanih tekmah.
    """
    preostale = set()
    if TEKME_FILE and os.path.exists(TEKME_FILE):
        try:
            wb = openpyxl.load_workbook(TEKME_FILE, data_only=True)
            ws = wb.worksheets[IZLOCILNI_SHEET_IDX]
            for row in ws.iter_rows(min_row=2, values_only=True):
                stage, ekipa1, ekipa2, datum, obdelana = (row + (None,)*5)[:5]
                if stage is None: break
                if str(obdelana or "").upper() == "D":
                    continue
                if ekipa1: preostale.add(str(ekipa1).strip().upper())
                if ekipa2: preostale.add(str(ekipa2).strip().upper())
            wb.close()
        except Exception:
            pass
    return preostale

def igralec_je_izpadel(napoved_path, ime_cell, preostale_ekipe, tekme_neobdelane):
    """
    Preveri, ali ima igralec v svoji napovedi (datoteka napoved_path) vsaj eno ekipo,
    ki je še med 'preostale_ekipe' (torej še lahko prinese točke).
    Vrne True če je igralec izpadel (nobena njegova napovedana ekipa ni več aktivna).
    tekme_neobdelane: seznam tekem (iz preberi_izlocilne_iz_excela), že filtriran na neobdelane.
    """
    if not preostale_ekipe:
        return False  # ni podatkov o tekmah – ne moremo soditi, privzeto ne pobarvaj
    try:
        wb = openpyxl.load_workbook(napoved_path, data_only=True)
        ws = wb.worksheets[0]
        kandidat_najden = False
        for t in tekme_neobdelane:
            for polje in (t["polje_e1"], t["polje_e2"]):
                if not polje: continue
                col, row = razdeli_celico(polje)
                if col is None: continue
                celica = ws[f"{col}{row}"].value
                if celica:
                    kandidat_najden = True
                    if str(celica).strip().upper() in preostale_ekipe:
                        wb.close()
                        return False  # ima vsaj eno aktivno ekipo
        wb.close()
        return kandidat_najden  # izpadel, samo če smo sploh kaj napovedi prebrali
    except Exception:
        return False

def sortiraj_in_ostevilci(rezultati_file=None, napovedi_mapa=None,
                            ime_cell=None, preveri_izpadle=False, log_cb=None,
                            start_row=None):
    """
    Sortira in oštevilči lestvico v podanem excelu rezultatov.
    Privzeto deluje na predtekmovanju (REZULTATI_FILE, start_row=3); za izločilni del podaj
    rezultati_file=REZULTATI_FILE_IZL, napovedi_mapa=NAPOVEDI_MAPA_IZL,
    ime_cell=IME_CELL_IZL, preveri_izpadle=True, start_row=4.
    """
    target_file = rezultati_file or REZULTATI_FILE
    sr = start_row if start_row is not None else 3
    if log_cb is None:
        log_cb = lambda m: None
    app = xw.App(visible=False)
    try:
        wb       = app.books.open(target_file)
        ws       = wb.sheets[0]
        last_row = ws.used_range.last_cell.row
        last_col = ws.used_range.last_cell.column
        # Izračunaj vsoto stolpcev D naprej in zapiši v C (zamenja =SUM formulo)
        for row in range(sr, last_row + 1):
            if ws.range((row, 2)).value is None: continue
            vrednosti = ws.range((row, 4), (row, last_col)).value
            if isinstance(vrednosti, (int, float)):
                vrednosti = [vrednosti]
            vsota = sum(v for v in (vrednosti or []) if isinstance(v, (int, float)))
            ws.range((row, 3)).value = vsota

        igralci  = []
        for row in range(sr, last_row + 1):
            vr = ws.range((row, 1), (row, last_col)).value
            if vr[1] is None: continue
            igralci.append(vr)
        igralci.sort(key=lambda x: x[2] if isinstance(x[2], (int, float)) else 0, reverse=True)
        for idx, vr in enumerate(igralci, start=sr):
            ws.range((idx, 1), (idx, last_col)).value = vr
        trenutno = 1
        prej     = None
        for row in range(sr, len(igralci) + sr):
            tocke = ws.range((row, 3)).value
            if row == sr:
                ws.range((row, 1)).value = "1."
            else:
                if tocke == prej:
                    ws.range((row, 1)).value = ""
                else:
                    trenutno = row - (sr - 1)
                    ws.range((row, 1)).value = f"{trenutno}."
            prej = tocke

        # Pobarvaj imena igralcev, ki ne morejo več doseči točk (vse njihove ekipe izpadle)
        if preveri_izpadle and napovedi_mapa and os.path.exists(napovedi_mapa):
            preostale_ekipe = preberi_aktivne_in_izpadle_ekipe()
            tekme_neobdelane = [t for t in preberi_izlocilne_iz_excela()
                                if str(t["obdelana"] or "").upper() != "D"]
            files = {f.lower(): f for f in os.listdir(napovedi_mapa)
                     if f.endswith(".xlsx") and not f.startswith("~")}
            stevec_izpadlih = 0
            for row in range(sr, len(igralci) + sr):
                ime = ws.range((row, 2)).value
                if not ime: continue
                ime_norm = str(ime).strip().title()
                # Poišči datoteko po imenu igralca
                ujemanje = None
                for fname in files.values():
                    try:
                        wbn = openpyxl.load_workbook(os.path.join(napovedi_mapa, fname),
                                                       read_only=True, data_only=True)
                        ime_v_datoteki = wbn.worksheets[0][ime_cell or IME_CELL_IZL].value
                        wbn.close()
                        if ime_v_datoteki and str(ime_v_datoteki).strip().title() == ime_norm:
                            ujemanje = fname
                            break
                    except Exception:
                        continue
                izpadel = False
                if ujemanje:
                    path = os.path.join(napovedi_mapa, ujemanje)
                    izpadel = igralec_je_izpadel(path, ime_cell or IME_CELL_IZL,
                                                   preostale_ekipe, tekme_neobdelane)
                barva = (220, 38, 38) if izpadel else (0, 0, 0)  # rdeča / črna
                ws.range((row, 2)).font.color = barva
                if izpadel:
                    stevec_izpadlih += 1
            log_cb(f"  {stevec_izpadlih} igralcev brez možnosti za nove točke (obarvani rdeče).")

        wb.save(target_file)
        wb.close()
        log_cb("✓  Lestvica uspešno sortirana in oštevilčena.")
    except Exception as e:
        log_cb(f"✗  Napaka: {e}")
    finally:
        app.quit()
    if not preveri_izpadle:
        IMENA_VRSTICE.clear()
        napolni_slovar_imen()
    else:
        IMENA_VRSTICE_IZL.clear()
        napolni_slovar_imen_izl()

def preberi_tekme_iz_excela():
    """Vrne seznam dict-ov iz lista Tekme v TEKME_FILE (vključno s številko vrstice)."""
    if not TEKME_FILE or not os.path.exists(TEKME_FILE):
        return []
    wb  = openpyxl.load_workbook(TEKME_FILE, read_only=True, data_only=True)
    ws  = wb.worksheets[TEKME_SHEET_IDX]
    tekme = []
    MESECI = {"jan":1,"feb":2,"mar":3,"apr":4,"maj":5,"jun":6,
               "jul":7,"avg":8,"sep":9,"okt":10,"nov":11,"dec":12,
               "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
               "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

    def format_datum(datum):
        if not datum: return ""
        try:
            if hasattr(datum, "day"):  # datetime objekt
                return f"{datum.day}.{datum.strftime('%b')}"
            return str(datum).strip()
        except Exception:
            return str(datum).strip()

    def datum_kljuc(datum):
        if not datum: return (99, 99)
        try:
            if hasattr(datum, "month"):  # datetime objekt
                return (datum.month, datum.day)
            delen = str(datum).strip().replace(".", " ").split()
            dan = int(delen[0])
            mes = delen[1].lower()[:3]
            return (MESECI.get(mes, 99), dan)
        except Exception:
            return (99, 99)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        ekipa1, sep, ekipa2, obdelana, stolpec, vrstica_napovedi, datum = (row + (None,)*7)[:7]
        if ekipa1 is None: break
        tekme.append({"ekipa1": ekipa1, "ekipa2": ekipa2,
                       "obdelana": obdelana, "stolpec": stolpec,
                       "vrstica_napovedi": vrstica_napovedi,
                       "datum": datum,
                       "datum_str": format_datum(datum),
                       "vrstica_excela": row_idx})
    wb.close()
    tekme.sort(key=lambda t: datum_kljuc(t["datum"]))
    return tekme

def oznaci_tekmo_obdelano(vrstica_excela):
    """Zapiše 'D' v stolpec obdelana (D) za dano vrstico v TEKME_FILE."""
    if not TEKME_FILE or not os.path.exists(TEKME_FILE):
        return
    wb = openpyxl.load_workbook(TEKME_FILE)
    ws = wb.worksheets[TEKME_SHEET_IDX]
    # Stolpec D (indeks 4) je stolpec "obdelana"
    ws.cell(row=vrstica_excela, column=4).value = "D"
    wb.save(TEKME_FILE)
    wb.close()


def oznaci_skupino_obdelano(vrstica_excela):
    """Zapiše 'D' v stolpec obdelana (E) za dano skupino v TEKME_FILE."""
    if not TEKME_FILE or not os.path.exists(TEKME_FILE):
        return
    wb = openpyxl.load_workbook(TEKME_FILE)
    ws = wb.worksheets[SKUPINE_SHEET_IDX]
    ws.cell(row=vrstica_excela, column=5).value = "D"
    wb.save(TEKME_FILE)
    wb.close()

def preveri_rezultati_prazni():
    """
    Vrne (True, "") če je excel rezultatov prazen (nima imen ali so vse točke 0).
    Vrne (False, razlog) sicer.
    """
    if not os.path.exists(REZULTATI_FILE):
        return True, ""
    try:
        app = xw.App(visible=False)
        try:
            wb  = app.books.open(REZULTATI_FILE)
            ws  = wb.sheets[0]
            # Preveri, ali stolpec B sploh vsebuje kakšno ime
            ime_celica = ws.range(f"{IME_COLUMN}3").value
            if ime_celica is None:
                return True, ""
            # Preveri vsoto stolpca C od vrstice 3 naprej
            vrednosti_c = ws.range("C3").expand("down").value
            wb.close()
        finally:
            app.quit()
        if vrednosti_c is None:
            return False, "Stolpec B že vsebuje imena, točke so prazne."
        if isinstance(vrednosti_c, (int, float)):
            vrednosti_c = [vrednosti_c]
        vsota = sum(v for v in vrednosti_c if isinstance(v, (int, float)))
        if vsota != 0:
            return False, f"Stolpec B že vsebuje imena in vsota točk (stolpec C) je {vsota:.0f} ≠ 0."
        return False, "Stolpec B že vsebuje imena (točke so še 0)."
    except Exception as e:
        return False, f"Napaka pri preverjanju: {e}"

def preberi_skupine_iz_excela():
    """Vrne seznam dict-ov iz lista Skupine v TEKME_FILE."""
    if not TEKME_FILE or not os.path.exists(TEKME_FILE):
        return []
    wb  = openpyxl.load_workbook(TEKME_FILE, read_only=True, data_only=True)
    ws  = wb.worksheets[SKUPINE_SHEET_IDX]
    skupine = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        skupina, stolpec, nap_stolpec, nap_vrstica, obdelana = (row + (None,)*5)[:5]
        if skupina is None: break
        skupine.append({"skupina": skupina, "stolpec": stolpec,
                        "nap_stolpec": nap_stolpec, "nap_vrstica": nap_vrstica,
                        "obdelana": obdelana, "vrstica_excela": row_idx})
    wb.close()
    return skupine

# ──────────────────────────────────────────────
#  IZLOČILNI DEL
# ──────────────────────────────────────────────

def napolni_slovar_imen_izl():
    """Slovar imen igralcev -> vrstica, za excel rezultatov izločilnih."""
    global IMENA_VRSTICE_IZL
    app = xw.App(visible=False)
    try:
        wb  = app.books.open(REZULTATI_FILE_IZL)
        ws  = wb.sheets[0]
        imena = ws.range(f"{IME_COLUMN_IZL}4").expand("down").value
        if imena:
            if not isinstance(imena, list):
                imena = [imena]
            IMENA_VRSTICE_IZL = {ime.strip().title(): i for i, ime in enumerate(imena, start=4) if ime}
        wb.close()
    finally:
        app.quit()

def prenesi_v_izlocilne(log_cb):
    """Prepiše imena in skupne točke iz REZULTATI_FILE (predtekmovanje) v REZULTATI_FILE_IZL."""
    app = xw.App(visible=False)
    try:
        wb_pred  = app.books.open(REZULTATI_FILE)
        ws_pred  = wb_pred.sheets[0]
        last_row = ws_pred.used_range.last_cell.row

        podatki = []
        for row in range(3, last_row + 1):
            mesto = ws_pred.range((row, 1)).value
            ime   = ws_pred.range((row, 2)).value
            tocke = ws_pred.range((row, 3)).value
            if ime is None: continue
            podatki.append((mesto, ime, tocke))
        wb_pred.close()

        wb_izl = app.books.open(REZULTATI_FILE_IZL)
        ws_izl = wb_izl.sheets[0]
        for i, (mesto, ime, tocke) in enumerate(podatki, start=4):
            ws_izl.range((i, 1)).value = mesto
            ws_izl.range((i, 2)).value = ime
            ws_izl.range((i, 4)).value = tocke  # stolpec D = predtekmovanje
            log_cb(f"✓  {ime}: {tocke} točk (predtekmovanje)")
        wb_izl.save(REZULTATI_FILE_IZL)
        wb_izl.close()
    finally:
        app.quit()
    log_cb(f"\n── Povzetek ──────────────────────")
    log_cb(f"Preneseno {len(podatki)} igralcev v izločilni del.")
    napolni_slovar_imen_izl()

def preberi_izlocilne_iz_excela():
    """Vrne seznam dict-ov iz lista Izlocilni v TEKME_FILE."""
    if not TEKME_FILE or not os.path.exists(TEKME_FILE):
        return []
    wb = openpyxl.load_workbook(TEKME_FILE, data_only=True)
    ws = wb.worksheets[IZLOCILNI_SHEET_IDX]
    tekme = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        stage, ekipa1, ekipa2, datum, obdelana, polje_e1, polje_e2, polje_nap, st_tock, stolpec_rez = (row + (None,)*10)[:10]
        if stage is None: break
        tekme.append({
            "stage": stage, "ekipa1": ekipa1, "ekipa2": ekipa2,
            "datum": datum, "obdelana": obdelana,
            "polje_e1": polje_e1, "polje_e2": polje_e2,
            "polje_nap": polje_nap, "st_tock": st_tock,
            "stolpec_rez": stolpec_rez, "vrstica_excela": row_idx,
        })
    wb.close()
    return tekme

def razdeli_celico(ref):
    """'B9' -> ('B', 9)"""
    import re
    m = re.match(r"([A-Za-z]+)(\d+)", str(ref).strip())
    if not m: return None, None
    return m.group(1).upper(), int(m.group(2))

def stolpec_gola(ws, col, row):
    """
    Vrne pravilen stolpec (kot črka) za gol, ki je takoj desno od celice imena
    (col, row). Če je celica imena del merge obsega (npr. K19:L22), gol ni v
    col+1 (kar bi padlo znotraj istega merge-a in vrnilo None), ampak v stolpcu
    takoj za KONCEM tega merge obsega.
    """
    col_idx = column_index_from_string(col)
    konec_col_idx = col_idx
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= col_idx <= mc.max_col:
            konec_col_idx = mc.max_col
            break
    return get_column_letter(konec_col_idx + 1)

def mozne_ekipe_za_slot(celica, tekme, obiskane=None):
    """
    Rekurzivno poišče vse ekipe, ki lahko teoretično pristanejo v dani celici
    (npr. 'E11'), sledeč verigi 'polje_nap' -> 'polje_e1'/'polje_e2' nazaj do R16,
    kjer so Ekipa1/Ekipa2 že znana imena.

    Pomembno: za R16 tekme je Ekipa1/Ekipa2 znana VNAPREJ (pred kakršnokoli obdelavo) –
    to je edina možna ekipa za ta slot. Za vse kasnejše faze (R8, R4, R2, FIN) postane
    Ekipa1/Ekipa2 znana šele PO obdelavi prejšnjega kroga in predstavlja dejanski
    rezultat napovedovanja, ne nabor možnih napovedi – zato moramo za te faze VEDNO
    slediti verigi nazaj do R16, ne glede na to, ali je Ekipa1/Ekipa2 že zapisana.

    celica: oznaka celice kot string (npr. 'E11') ali že znano ime ekipe.
    tekme: seznam vseh tekem iz preberi_izlocilne_iz_excela().
    Vrne množico imen ekip (kot so zapisane v Ekipa1/Ekipa2, originalna velikost črk).
    """
    if obiskane is None:
        obiskane = set()
    if not celica:
        return set()
    celica_norm = str(celica).strip().upper()
    if celica_norm in obiskane:
        return set()  # varovalo pred ciklom
    obiskane.add(celica_norm)

    # Če celica ni oblike "črka+številka" (npr. je že ime ekipe), vrni jo kot tako
    col, row = razdeli_celico(celica_norm)
    if col is None:
        return {str(celica).strip()}

    # Poišči tekmo, kjer se celica ujema z polje_e1 ali polje_e2 (znan slot te faze)
    for t in tekme:
        je_r16 = str(t["stage"]).strip().upper() == "R16"
        if t["polje_e1"] and str(t["polje_e1"]).strip().upper() == celica_norm:
            if je_r16 and t["ekipa1"]:  # R16 – ekipa znana vnaprej, edina možnost
                return {str(t["ekipa1"]).strip()}
            # R8+ – ne glede na to, ali je Ekipa1 že zapisana (dejanski rezultat),
            # kandidati za validacijo izhajajo iz izvornih R16 ekip
            return _ekipe_prejsnje_tekme(celica_norm, tekme, obiskane)
        if t["polje_e2"] and str(t["polje_e2"]).strip().upper() == celica_norm:
            if je_r16 and t["ekipa2"]:
                return {str(t["ekipa2"]).strip()}
            return _ekipe_prejsnje_tekme(celica_norm, tekme, obiskane)

    # Ni najdene tekme s tem slotom – morda je to neposredno ime ekipe
    return {str(celica).strip()}

def _ekipe_prejsnje_tekme(celica_norm, tekme, obiskane):
    """
    Poišče tekmo, katere 'polje_nap' je enak celica_norm, in vrne unijo možnih ekip
    obeh strani te tekme (rekurzivno).
    """
    for t in tekme:
        if t["polje_nap"] and str(t["polje_nap"]).strip().upper() == celica_norm:
            ekipe = set()
            ekipe |= mozne_ekipe_za_slot(t["polje_e1"], tekme, obiskane)
            ekipe |= mozne_ekipe_za_slot(t["polje_e2"], tekme, obiskane)
            return ekipe
    return set()

def zapisi_napredujočo_ekipo(vrstica_excela, ime_ekipe, log_cb):
    """
    Po obdelavi tekme zapiše napovedano napredujočo ekipo v ustrezno polje
    (Ekipa1/Ekipa2) naslednje tekme v TEKME_FILE, glede na ujemanje oznake celice
    'polje_nap' te tekme s 'polje_e1'/'polje_e2' druge tekme.
    """
    if not TEKME_FILE or not os.path.exists(TEKME_FILE):
        return
    wb = openpyxl.load_workbook(TEKME_FILE)
    ws = wb.worksheets[IZLOCILNI_SHEET_IDX]

    polje_nap = ws.cell(row=vrstica_excela, column=8).value  # H = Polje napredujoče ekipe
    if not polje_nap:
        wb.close()
        return

    polje_nap = str(polje_nap).strip().upper()
    zapisano = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        polje_e1 = row[5].value  # F = Polje ekipe 1
        polje_e2 = row[6].value  # G = Polje ekipe 2
        if polje_e1 and str(polje_e1).strip().upper() == polje_nap:
            ws.cell(row=row[0].row, column=2).value = ime_ekipe  # B = Ekipa1
            zapisano = True
            log_cb(f"  → {ime_ekipe} zapisana kot Ekipa1 v vrstici {row[0].row}")
            break
        if polje_e2 and str(polje_e2).strip().upper() == polje_nap:
            ws.cell(row=row[0].row, column=3).value = ime_ekipe  # C = Ekipa2
            zapisano = True
            log_cb(f"  → {ime_ekipe} zapisana kot Ekipa2 v vrstici {row[0].row}")
            break

    wb.save(TEKME_FILE)
    wb.close()
    if not zapisano:
        log_cb(f"  ⚠  Ni najdene naslednje tekme za polje {polje_nap} (morda finale/3.mesto že odigrano).")

def oznaci_izlocilno_obdelano(vrstica_excela):
    """Zapiše 'D' v stolpec 'Obdelana tekma' (E) za dano tekmo v TEKME_FILE."""
    if not TEKME_FILE or not os.path.exists(TEKME_FILE):
        return
    wb = openpyxl.load_workbook(TEKME_FILE)
    ws = wb.worksheets[IZLOCILNI_SHEET_IDX]
    ws.cell(row=vrstica_excela, column=5).value = "D"
    wb.save(TEKME_FILE)
    wb.close()

def obdelaj_izlocilno_tekmo(tekma, gd, gg, napredujoca_ekipa, log_cb, progress_cb=None):
    """
    Obdela eno izločilno tekmo za vse igralce.
    tekma: dict iz preberi_izlocilne_iz_excela()
    gd, gg: dejanski rezultat (90 min)
    napredujoca_ekipa: dejanska ekipa ki gre naprej (lahko = ekipa z več goli, ali ob izenačenju ročno izbrana)
    """
    files = [f for f in os.listdir(NAPOVEDI_MAPA_IZL) if f.endswith(".xlsx") and not f.startswith("~")]
    skupaj = len(files)
    stat = {"tip": 0, "rezultat": 0, "napredovanje": 0, "nic": 0, "napake": 0}

    col_e1, row_e1 = razdeli_celico(tekma["polje_e1"])
    col_e2, row_e2 = razdeli_celico(tekma["polje_e2"])
    col_nap, row_nap = razdeli_celico(tekma["polje_nap"])
    col_rez_tip  = tekma["stolpec_rez"]
    col_rez_nap  = get_column_letter(column_index_from_string(col_rez_tip) + 1)
    st_tock_nap  = tekma["st_tock"]

    ekipa1, ekipa2 = tekma["ekipa1"], tekma["ekipa2"]
    ekipe_originalne = [ekipa1, ekipa2]
    ekipe_polne = [str(e).upper() for e in ekipe_originalne]

    tip_rez = vrni_tip(gd, gg)

    app = xw.App(visible=False)
    try:
        wb = app.books.open(REZULTATI_FILE_IZL)
        ws = wb.sheets[0]

        for i, file in enumerate(files):
            path = os.path.join(NAPOVEDI_MAPA_IZL, file)
            try:
                wbn = openpyxl.load_workbook(path, data_only=True)
                wsn = wbn.worksheets[0]
                ime = wsn[IME_CELL_IZL].value.strip().title()
                vr  = IMENA_VRSTICE_IZL.get(ime)
                if vr is None: raise Exception("Ime ni najdeno!")

                # Preberi napoved imen ekip
                celica_e1 = wsn[f"{col_e1}{row_e1}"].value
                celica_e2 = wsn[f"{col_e2}{row_e2}"].value
                napoved_e1 = str(celica_e1).strip() if celica_e1 else ""
                napoved_e2 = str(celica_e2).strip() if celica_e2 else ""

                # Preberi napovedan rezultat (gol je v stolpcu desno od KONCA merge obsega imena)
                gol1_col = stolpec_gola(wsn, col_e1, row_e1)
                gol2_col = stolpec_gola(wsn, col_e2, row_e2)
                gol1 = wsn[f"{gol1_col}{row_e1}"].value
                gol2 = wsn[f"{gol2_col}{row_e2}"].value

                if str(tekma["stage"]).strip().upper() == "3M":
                    # Tekma za 3. mesto: vrstni red zgornja/spodnja ekipa ni intuitiven,
                    # zato preverimo obe možni razporeditvi in vzamemo boljšo.
                    match_e1, proc_e1 = razreši_z_aliasi(napoved_e1, ekipe_polne, ekipe_originalne) if napoved_e1 else (None, 0)
                    match_e2, proc_e2 = razreši_z_aliasi(napoved_e2, ekipe_polne, ekipe_originalne) if napoved_e2 else (None, 0)

                    par_a_ok = (proc_e1 >= PRAG_VALIDACIJE and proc_e2 >= PRAG_VALIDACIJE and
                                match_e1 == ekipe_polne[0] and match_e2 == ekipe_polne[1])
                    par_b_ok = (proc_e1 >= PRAG_VALIDACIJE and proc_e2 >= PRAG_VALIDACIJE and
                                match_e1 == ekipe_polne[1] and match_e2 == ekipe_polne[0])

                    if par_a_ok:
                        pravilen_par = True
                        gol1_uskl, gol2_uskl = gol1, gol2  # zaporedje že ustreza ekipa1:ekipa2
                    elif par_b_ok:
                        pravilen_par = True
                        gol1_uskl, gol2_uskl = gol2, gol1  # igralec je napovedal obratno, zamenjamo gole
                    else:
                        pravilen_par = False
                        gol1_uskl, gol2_uskl = gol1, gol2
                else:
                    match_e1, proc_e1 = razreši_z_aliasi(napoved_e1, ekipe_polne, ekipe_originalne) if napoved_e1 else (None, 0)
                    match_e2, proc_e2 = razreši_z_aliasi(napoved_e2, ekipe_polne, ekipe_originalne) if napoved_e2 else (None, 0)
                    pravilen_par = (proc_e1 >= PRAG_VALIDACIJE and proc_e2 >= PRAG_VALIDACIJE and
                                    match_e1 == ekipe_polne[0] and match_e2 == ekipe_polne[1])
                    gol1_uskl, gol2_uskl = gol1, gol2

                tocke_tekma = 0
                razlog_nic = None
                if pravilen_par:
                    if gol1_uskl is None or gol2_uskl is None:
                        stat["nic"] += 1
                        razlog_nic = "manjka rezultat"
                    else:
                        try:
                            gol1n, gol2n = int(gol1_uskl), int(gol2_uskl)
                            if gol1n == gd and gol2n == gg:
                                tocke_tekma = 6
                                stat["rezultat"] += 1
                            elif vrni_tip(gol1n, gol2n) == tip_rez:
                                tocke_tekma = 3
                                stat["tip"] += 1
                            else:
                                stat["nic"] += 1
                        except (TypeError, ValueError):
                            stat["nic"] += 1
                            razlog_nic = "neveljaven rezultat"
                else:
                    stat["nic"] += 1
                    razlog_nic = "manjka/napačno ime ekipe"

                ws.range(f"{col_rez_tip}{vr}").value = tocke_tekma

                # Napovedana napredujoča ekipa
                celica_nap = wsn[f"{col_nap}{row_nap}"].value
                napoved_nap = str(celica_nap).strip() if celica_nap else ""
                tocke_nap = 0
                if napoved_nap:
                    match_nap, proc_nap = razreši_z_aliasi(napoved_nap, ekipe_polne, ekipe_originalne)
                    if proc_nap >= PRAG_VALIDACIJE and match_nap == str(napredujoca_ekipa).upper():
                        tocke_nap = st_tock_nap
                        stat["napredovanje"] += 1
                ws.range(f"{col_rez_nap}{vr}").value = tocke_nap

                wbn.close()
                logging.info(f"{ime} | {tekma['stage']} {ekipa1}:{ekipa2} | tekma={tocke_tekma}, napr={tocke_nap}")
                dodatek = f"  ({razlog_nic})" if razlog_nic else ""
                log_cb(f"✓  {ime}: {tocke_tekma} (tekma) + {tocke_nap} (napredovanje){dodatek}")
            except Exception as e:
                stat["napake"] += 1
                logging.error(f"NAPAKA {file}: {e}")
                log_cb(f"✗  {file}: {e}")
            if progress_cb: progress_cb(int((i + 1) / skupaj * 100))

        wb.save(REZULTATI_FILE_IZL)
        wb.close()
    finally:
        app.quit()

    log_cb(f"\n── Povzetek ──────────────────────")
    log_cb(f"Obdelanih: {skupaj - stat['napake']}/{skupaj}  |  "
           f"6t (rezultat): {stat['rezultat']}  |  3t (tip): {stat['tip']}  |  "
           f"0t: {stat['nic']}  |  Napredovanje: {stat['napredovanje']}  |  Napake: {stat['napake']}")

# ──────────────────────────────────────────────
#  VALIDACIJA NAPOVEDI SKUPIN
# ──────────────────────────────────────────────

SKUPINE_SP2026 = {
    "A": ["Mehika", "Južna Afrika", "Južna Koreja", "Češka"],
    "B": ["Kanada", "Bosna in Hercegovina", "Katar", "Švica"],
    "C": ["Brazilija", "Maroko", "Škotska", "Haiti"],
    "D": ["Združene države Amerike", "Paragvaj", "Avstralija", "Turčija"],
    "E": ["Nemčija", "Slonokoščena obala", "Ekvador", "Curacao"],
    "F": ["Nizozemska", "Japonska", "Švedska", "Tunizija"],
    "G": ["Belgija", "Egipt", "Iran", "Nova Zelandija"],
    "H": ["Španija", "Urugvaj", "Savdska Arabija", "Zelenortski otoki"],
    "I": ["Francija", "Senegal", "Irak", "Norveška"],
    "J": ["Argentina", "Alžirija", "Avstrija", "Jordan"],
    "K": ["Portugalska", "DR Kongo", "Uzbekistan", "Kolumbija"],
    "L": ["Anglija", "Hrvaška", "Gana", "Panama"],
}

# Alternativna imena za ekipe (kratice, okrajšave, pogovorni izrazi)
# Ključ je kanonično ime (mora biti enako kot v SKUPINE_SP2026), vrednost seznam aliasov
ALIASI_EKIP = {
    "BOSNA IN HERCEGOVINA": ["BIH", "BOSNA"],
    "ZDRUŽENE DRŽAVE AMERIKE": ["ZDA", "USA", "AMERIKA"],
    "SLONOKOŠČENA OBALA": ["SLONOK.O .", "SLONOKO. O.", "IVORY COAST", "SLONOKOŠČENA O.", "CÔTE D'IVOIRE", "COTE D'IVOIRE", "SLONOK. OBALA", "SLONOKOSCEN O.", "SL.OBALA", "SLONOKOSCENA O.", "SLONOK. O."],
    "JUŽNA KOREJA": ["KOREJA", "KOREA", "J KOREJA", "J. KOREJA", "J.KOREJA", "SOUTH KOREA", "J. KOREA"],
    "JUŽNA AFRIKA": ["JAR", "RSA", "J. AFRIKA", "J AFRIKA", "J.AFRIKA", "JUŽNOAFRIŠKA REPUBLIKA", "SOUTH AFRICA", "J. AFRKA"],
    "NOVA ZELANDIJA": ["N ZELANDIJA", "N. ZELANDIJA", "N.ZELANDIJA", "N.ZELENDIJA", "N. ZELENDIJA"],
    "SAVDSKA ARABIJA": ["SAU.ARABIJA", "SAV. ARABIA", "S. ARABIJA", "SAV. ARABIJA", "SAV.ARABIJA", "S ARABIJA", "S.ARABIJA", "SAV ARABIJA", "SAU ARABIJA", "SA.ARABIJA", "SAUDI ARABIA", "SAU. ARABIJA"],
    "ZELENORTSKI OTOKI": ["Z.OSTRVA", "ZELE.OTOKI", "KAPVERDSKI O.", "KAPVERD.OTOKI", "ZELENO. OTO.", "ZELENORTSKI O.", "Z. OTOKI", "CAPE VERDE", "ZELENORTSKI.O", "ZELENO R OTOKI", "ZELENOERTSKI", "KAPVERDSKI OTOKI", "ZEL.OTOKI", "ZELENORTSKI.O."],
    "DR KONGO": ["DRC", "CONGO DR", "D.R. KONGO", "D.R.KONGO", "D. R. KONGO", "D.R.", "CONGO", "KONGO"],
    "CURACAO": ["CURAÇAO"],
    "FRANCIJA": ["FRANCOSKA"],
    "PORTUGALSKA": ["PORTUGALIJA", "PORTUGAL"],
    "NEMČIJA": ["NEMCIJA"],
    "ČEŠKA": ["CESKA", "ČESKA"],
    "JORDAN": ["JORDANIJA"],
    "ALŽIRIJA": ["ALĐERIJA", "ALZ"],
    "ŠVICA": ["SVICA"],
    "KATAR": ["QATAR"],
    "KANADA": ["CANADA"],
    "MAROKO": ["MAROCO"],
    "EGIPT": ["EGYPT"],
    "AVSTRIJA": ["AUSTRIA"],
    "HRVAŠKA": ["CROATIA"],
    "IRAK": ["IRAQ"],
    "ŠKOTSKA": ["SCOTLAND"],
    "NIZOZEMSKA": ["NIZOZEMCI"],
    "BIH": ["BOSNA"],
    "J. AFRIKA": ["JAR"],
    "EKVADOR": ["ECUADOR"]
}

def razreši_z_aliasi(vnos, ekipe_polna, ekipe_originalne):
    """
    Najprej preveri aliase – če vnos direktno ustreza aliasu, vrni 100%.
    Sicer vrni najboljši fuzzy match.
    ekipe_polna: seznam UPPER imen (za fuzzy)
    ekipe_originalne: seznam originalnih imen (za alias lookup)
    """
    vnos_upper = vnos.strip().upper()

    # Preveri aliase za vsako ekipo
    for i, ime_orig in enumerate(ekipe_originalne):
        aliasi = ALIASI_EKIP.get(ime_orig.upper(), [])
        for alias in aliasi:
            if vnos_upper == alias.upper():
                return ekipe_polna[i], 100.0

    # Fallback na fuzzy
    match, procent, _ = process.extractOne(vnos_upper, ekipe_polna)
    return match, float(procent)


PRAG_VALIDACIJE = 75  # minimalni % ujemanja

def validiraj_napovedi_skupin(skupine_ekipe, log_cb):
    """
    skupine_ekipe: seznam dict-ov {"skupina": "A", "stolpec": "X", "nap_stolpec": "Y",
                                    "nap_vrstica": N, "ekipe": ["Ekipa1", ...]}
    Gre čez vse napovedi in preverja ujemanje imen ekip v skupinah.
    """
    files = [f for f in os.listdir(NAPOVEDI_MAPA) if f.endswith(".xlsx") and not f.startswith("~")]
    problemi = []
    skupaj_datotek = len(files)
    log_cb(f"Preverjam {skupaj_datotek} napovedi …\n")

    for file in files:
        path = os.path.join(NAPOVEDI_MAPA, file)
        problemi_datoteke = []
        try:
            wbn = openpyxl.load_workbook(path, data_only=True)
            wsn = wbn.worksheets[0]
            ime = wsn[IME_CELL].value.strip().title()

            for sk in skupine_ekipe:
                nst = str(sk["nap_stolpec"]).strip().upper()
                nvr = int(sk["nap_vrstica"])
                ekipe_originalne = [e.strip() for e in sk["ekipe"]]
                ekipe_polne = [e.upper() for e in ekipe_originalne]

                for idx in range(4):
                    celica = wsn[f"{nst}{nvr + idx}"].value
                    if celica is None:
                        problemi_datoteke.append(
                            f"  Skupina {sk['skupina']}, mesto {idx+1}: celica prazna")
                        continue
                    vnos = str(celica).strip()
                    match, procent = razreši_z_aliasi(vnos, ekipe_polne, ekipe_originalne)
                    if procent < PRAG_VALIDACIJE:
                        problemi_datoteke.append(
                            f"  Skupina {sk['skupina']}, mesto {idx+1}: "
                            f"'{vnos}' → najbližje '{match.title()}' ({procent:.0f}%)")

            wbn.close()

            if problemi_datoteke:
                problemi.append((ime, file, problemi_datoteke))

        except Exception as e:
            log_cb(f"✗  {file}: {e}")

    if not problemi:
        log_cb("✓  Vse napovedi so v redu – ni problematičnih vnosov.")
    else:
        log_cb(f"⚠  Najdenih {len(problemi)} datotek s problemi:\n")
        for ime, file, pp in problemi:
            log_cb(f"▸  {ime}  ({file})")
            for p in pp:
                log_cb(p)
            log_cb("")

def validiraj_napovedi_izlocilnih(log_cb):
    """
    Gre čez vse napovedi izločilnega dela in za VSAKO še neobdelano tekmo
    (ne le tiste z že znano Ekipa1/Ekipa2) preveri:
    1. Popolnost vnosa (ime ekipe1, ime ekipe2, rezultat, napredujoča ekipa).
    2. Pravilnost vnesenih imen ekip – primerjano s kandidati, ki teoretično
       lahko pristanejo na tem mestu bracketa (rekurzivno nazaj do R16).
       Za 3. mesto (ni sledljivo prek polje_nap) se preveri samo popolnost.
    """
    vse_tekme = preberi_izlocilne_iz_excela()
    tekme = [t for t in vse_tekme if str(t["obdelana"] or "").upper() != "D"]

    if not tekme:
        log_cb("ℹ  Ni neobdelanih tekem za preverjanje.")
        return

    # Vnaprej izračunaj kandidate za vsak slot (polje_e1/polje_e2), da se izognemo
    # ponovnemu rekurzivnemu računanju za vsako napoved datoteko posebej.
    kandidati_cache = {}
    def kandidati_za(polje):
        if not polje:
            return set()
        key = str(polje).strip().upper()
        if key not in kandidati_cache:
            kandidati_cache[key] = mozne_ekipe_za_slot(polje, vse_tekme)
        return kandidati_cache[key]

    files = [f for f in os.listdir(NAPOVEDI_MAPA_IZL) if f.endswith(".xlsx") and not f.startswith("~")]
    problemi = []
    log_cb(f"Preverjam {len(files)} napovedi za {len(tekme)} neobdelanih tekem …\n")

    for file in files:
        path = os.path.join(NAPOVEDI_MAPA_IZL, file)
        problemi_datoteke = []
        try:
            wbn = openpyxl.load_workbook(path, data_only=True)
            wsn = wbn.worksheets[0]
            ime_cell_val = wsn[IME_CELL_IZL].value
            ime = ime_cell_val.strip().title() if ime_cell_val else file

            for t in tekme:
                naziv_tekme = (f"{t['stage']} {t['ekipa1']} : {t['ekipa2']}"
                               if t["ekipa1"] and t["ekipa2"] else f"{t['stage']} (vrstica {t['vrstica_excela']})")

                kand_e1 = kandidati_za(t["polje_e1"])
                kand_e2 = kandidati_za(t["polje_e2"])
                kand_oba = kand_e1 | kand_e2  # za napredujočo ekipo – lahko je iz katerekoli strani
                kand_oba_polna = [e.upper() for e in kand_oba]
                kand_oba_orig = list(kand_oba)

                col_e1, row_e1 = razdeli_celico(t["polje_e1"])
                col_e2, row_e2 = razdeli_celico(t["polje_e2"])
                col_nap, row_nap = razdeli_celico(t["polje_nap"]) if t["polje_nap"] else (None, None)

                # Ime ekipe 1 – popolnost + pravilnost (če imamo kandidate)
                e1 = wsn[f"{col_e1}{row_e1}"].value if col_e1 else None
                if not e1:
                    problemi_datoteke.append(f"  {naziv_tekme}: manjka napoved ekipe 1 ({t['polje_e1']})")
                elif kand_e1:
                    kand_e1_polna = [e.upper() for e in kand_e1]
                    match1, proc1 = razreši_z_aliasi(str(e1).strip(), kand_e1_polna, list(kand_e1))
                    if proc1 < PRAG_VALIDACIJE:
                        problemi_datoteke.append(
                            f"  {naziv_tekme}: napoved ekipe 1 '{e1}' → najbližje "
                            f"'{match1.title()}' ({proc1:.0f}%) ({t['polje_e1']})")

                # Ime ekipe 2 – popolnost + pravilnost (če imamo kandidate)
                e2 = wsn[f"{col_e2}{row_e2}"].value if col_e2 else None
                if not e2:
                    problemi_datoteke.append(f"  {naziv_tekme}: manjka napoved ekipe 2 ({t['polje_e2']})")
                elif kand_e2:
                    kand_e2_polna = [e.upper() for e in kand_e2]
                    match2, proc2 = razreši_z_aliasi(str(e2).strip(), kand_e2_polna, list(kand_e2))
                    if proc2 < PRAG_VALIDACIJE:
                        problemi_datoteke.append(
                            f"  {naziv_tekme}: napoved ekipe 2 '{e2}' → najbližje "
                            f"'{match2.title()}' ({proc2:.0f}%) ({t['polje_e2']})")

                # Rezultat (gol je v stolpcu desno od KONCA merge obsega imena)
                if col_e1:
                    gol1_col = stolpec_gola(wsn, col_e1, row_e1)
                    gol1 = wsn[f"{gol1_col}{row_e1}"].value
                    if gol1 is None:
                        problemi_datoteke.append(f"  {naziv_tekme}: manjka rezultat ekipe 1 ({gol1_col}{row_e1})")
                if col_e2:
                    gol2_col = stolpec_gola(wsn, col_e2, row_e2)
                    gol2 = wsn[f"{gol2_col}{row_e2}"].value
                    if gol2 is None:
                        problemi_datoteke.append(f"  {naziv_tekme}: manjka rezultat ekipe 2 ({gol2_col}{row_e2})")

                # Napredujoča ekipa – popolnost + pravilnost (mora ustrezati eni od kandidatov obeh strani)
                if col_nap:
                    nap = wsn[f"{col_nap}{row_nap}"].value
                    if not nap:
                        problemi_datoteke.append(
                            f"  {naziv_tekme}: manjka napoved napredujoče ekipe ({t['polje_nap']})")
                    elif kand_oba:
                        match_nap, proc_nap = razreši_z_aliasi(str(nap).strip(), kand_oba_polna, kand_oba_orig)
                        if proc_nap < PRAG_VALIDACIJE:
                            problemi_datoteke.append(
                                f"  {naziv_tekme}: napredujoča ekipa '{nap}' → najbližje "
                                f"'{match_nap.title()}' ({proc_nap:.0f}%) ({t['polje_nap']})")

            wbn.close()

            if problemi_datoteke:
                problemi.append((ime, file, problemi_datoteke))

        except Exception as e:
            log_cb(f"✗  {file}: {e}")

    if not problemi:
        log_cb("✓  Vse napovedi so popolne in pravilne za vse neobdelane tekme.")
    else:
        log_cb(f"⚠  Najdenih {len(problemi)} datotek z manjkajočimi vnosi:\n")
        for ime, file, pp in problemi:
            log_cb(f"▸  {ime}  ({file})")
            for p in pp:
                log_cb(p)
            log_cb("")

# ──────────────────────────────────────────────
#  GUI
# ──────────────────────────────────────────────

DARK   = "#0f1117"
PANEL  = "#1a1d27"
CARD   = "#22263a"
ACCENT = "#3b82f6"
GREEN  = "#22c55e"
RED    = "#ef4444"
MUTED  = "#64748b"
FG     = "#e2e8f0"
FG2    = "#94a3b8"

FONT_BODY  = ("Inter", 10)
FONT_LABEL = ("Inter", 9)
FONT_MONO  = ("Consolas", 9)
FONT_HEAD  = ("Inter", 13, "bold")
FONT_TITLE = ("Inter", 18, "bold")


class SP2026App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("SP 2026 – Točkovanje")
        self.geometry("960x680")
        self.minsize(860, 600)
        self.configure(bg=DARK)
        self._build()
        self._refresh_globals()
        self.after(100, lambda: self._run_threaded(self._osveži_ob_zagonu))
        self.protocol('WM_DELETE_WINDOW', self._ob_zapiranju)

    # ── GRADNJA ──────────────────────────────

    def _ob_zapiranju(self):
        shrani_config()
        self.destroy()

    def _osveži_ob_zagonu(self):
        """Ob zagonu avtomatsko naloži tekme, skupine in slovarje imen v ozadju."""
        try:
            napolni_slovar_imen()
        except Exception:
            pass
        try:
            napolni_slovar_imen_izl()
        except Exception:
            pass
        if TEKME_FILE and os.path.exists(TEKME_FILE):
            self._osveži_tekme()
            self._osveži_skupine()
            self._osveži_izlocilne()

    def _build(self):
        # Leva navigacija
        nav = tk.Frame(self, bg=PANEL, width=190)
        nav.pack(side="left", fill="y")
        nav.pack_propagate(False)

        tk.Label(nav, text="⚽  SP 2026", bg=PANEL, fg=FG,
                 font=FONT_TITLE, pady=22).pack(fill="x", padx=16)
        tk.Frame(nav, bg=MUTED, height=1).pack(fill="x", padx=16, pady=4)

        self._frames = {}
        self._nav_btns = {}

        pages = [
            ("tekma",     "▷  Vnesi tekmo"),
            ("skupina",   "◈  Zaključi skupino"),
            ("imena",     "⊕  Naloži imena"),
            ("sortiraj",  "⇅  Sortiraj lestvico"),
            ("preveri",   "🔍  Preveri napovedi"),
            ("prenesi",   "⤴  Prenesi v izločilne"),
            ("izlocilna", "🏆  Vnesi izločilno"),
            ("nastavitve","⚙  Nastavitve"),
        ]

        content = tk.Frame(self, bg=DARK)
        content.pack(side="left", fill="both", expand=True)

        for key, label in pages:
            btn = tk.Button(
                nav, text=label, anchor="w",
                bg=PANEL, fg=FG2, bd=0,
                activebackground=CARD, activeforeground=FG,
                font=FONT_BODY, padx=20, pady=10,
                cursor="hand2",
                command=lambda k=key: self._show(k)
            )
            btn.pack(fill="x")
            self._nav_btns[key] = btn

        # Ustvari strani
        for key, _ in pages:
            f = tk.Frame(content, bg=DARK)
            f.place(relwidth=1, relheight=1)
            self._frames[key] = f

        self._build_tekma(self._frames["tekma"])
        self._build_skupina(self._frames["skupina"])
        self._build_imena(self._frames["imena"])
        self._build_sortiraj(self._frames["sortiraj"])
        self._build_nastavitve(self._frames["nastavitve"])
        self._build_preveri(self._frames["preveri"])
        self._build_prenesi(self._frames["prenesi"])
        self._build_izlocilna(self._frames["izlocilna"])

        self._show("tekma")

    def _show(self, key):
        self._frames[key].lift()
        for k, btn in self._nav_btns.items():
            btn.configure(bg=CARD if k == key else PANEL,
                          fg=FG  if k == key else FG2)

    # ── SKUPNE POMOŽNE METODE ─────────────────

    def _card(self, parent, title, row=0, col=0, colspan=1):
        f = tk.Frame(parent, bg=CARD, padx=18, pady=14)
        f.grid(row=row, column=col, columnspan=colspan,
               sticky="nsew", padx=8, pady=6)
        tk.Label(f, text=title, bg=CARD, fg=FG2,
                 font=FONT_LABEL).pack(anchor="w", pady=(0, 6))
        return f

    def _labeled_entry(self, parent, label, default="", width=22):
        tk.Label(parent, text=label, bg=CARD, fg=FG2,
                 font=FONT_LABEL).pack(anchor="w", pady=(6, 1))
        v = tk.StringVar(value=default)
        e = tk.Entry(parent, textvariable=v, bg=DARK, fg=FG,
                     insertbackground=FG, relief="flat",
                     font=FONT_BODY, width=width)
        e.pack(fill="x", ipady=5)
        return v

    def _log_box(self, parent, rows=10):
        txt = scrolledtext.ScrolledText(
            parent, height=rows, bg=DARK, fg=FG2,
            font=FONT_MONO, relief="flat",
            insertbackground=FG, state="disabled"
        )
        txt.pack(fill="both", expand=True, pady=(6, 0))
        return txt

    def _log(self, box, msg):
        box.configure(state="normal")
        box.insert("end", msg + "\n")
        box.see("end")
        box.configure(state="disabled")
        self.update_idletasks()

    def _btn(self, parent, text, cmd, color=ACCENT):
        return tk.Button(
            parent, text=text, command=cmd,
            bg=color, fg="white", relief="flat",
            font=("Inter", 10, "bold"),
            padx=14, pady=8, cursor="hand2",
            activebackground=DARK, activeforeground=FG
        )

    def _run_threaded(self, fn, *args, **kwargs):
        """
        Poženemo funkcijo direktno (brez ločenega threada).
        Excel COM (xlwings) se v praksi ne zažene zanesljivo iz ozadnjega
        threada na Windows, zato vse akcije tečejo sinhrono v glavnem GUI threadu.
        GUI se med izvajanjem ne odziva, kar je sprejemljivo glede na trajanje akcij.
        """
        self.update_idletasks()
        try:
            fn(*args, **kwargs)
        except Exception as e:
            logging.error(f"Napaka pri izvajanju {getattr(fn, '__name__', fn)}: {e}")
            raise

    # ── STRAN: TEKMA ──────────────────────────

    def _build_tekma(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Vnesi rezultat tekme",
                 bg=DARK, fg=FG, font=FONT_HEAD,
                 pady=16).grid(row=0, column=0, columnspan=2, sticky="w", padx=18)

        # Leva kartica – vnos
        left = self._card(parent, "Podatki tekme", row=1, col=0)

        # Dropdown iz excel datoteke
        tk.Label(left, text="Tekma iz excela (opcijsko)", bg=CARD,
                 fg=FG2, font=FONT_LABEL).pack(anchor="w", pady=(6, 1))
        self._tekma_var = tk.StringVar(value="— izberi —")
        self._tekma_combo = ttk.Combobox(
            left, textvariable=self._tekma_var,
            state="readonly", font=FONT_BODY, width=48
        )
        self._tekma_combo.pack(fill="x", ipady=4)
        self._tekma_combo.bind("<<ComboboxSelected>>", self._on_tekma_select)
        self._btn(left, "↻  Osveži seznam", self._osveži_tekme).pack(
            anchor="w", pady=(4, 10))

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=8)

        self._lbl_domaci = tk.Label(left, text="Goli domači", bg=CARD, fg=FG2, font=FONT_LABEL)
        self._lbl_domaci.pack(anchor="w", pady=(6, 1))
        self._t_domaci = tk.StringVar()
        tk.Entry(left, textvariable=self._t_domaci, bg=DARK, fg=FG,
                 insertbackground=FG, relief="flat", font=FONT_BODY, width=22).pack(fill="x", ipady=5)

        self._lbl_gostje = tk.Label(left, text="Goli gostje", bg=CARD, fg=FG2, font=FONT_LABEL)
        self._lbl_gostje.pack(anchor="w", pady=(6, 1))
        self._t_gostje = tk.StringVar()
        tk.Entry(left, textvariable=self._t_gostje, bg=DARK, fg=FG,
                 insertbackground=FG, relief="flat", font=FONT_BODY, width=22).pack(fill="x", ipady=5)

        self._btn(left, "▷  Obdelaj tekmo",
                  self._zacni_tekmo, GREEN).pack(anchor="w", pady=(14, 0))

        # Desna kartica – log
        right = self._card(parent, "Dnevnik", row=1, col=1)
        self._tekma_log = self._log_box(right, rows=16)
        self._tekma_progress = ttk.Progressbar(right, mode="determinate", maximum=100)
        self._tekma_progress.pack(fill="x", pady=(6, 0))

    def _osveži_tekme(self):
        vse_tekme = preberi_tekme_iz_excela()
        # Prikaži samo neobdelane (obdelana != "D")
        self._tekme_data = [t for t in vse_tekme if str(t["obdelana"] or "").upper() != "D"]
        vrednosti = [
            f"{t['datum_str']}  {t['ekipa1']} – {t['ekipa2']}" if t.get("datum_str") else f"{t['ekipa1']} – {t['ekipa2']}"
            for t in self._tekme_data
        ]
        self._tekma_combo["values"] = vrednosti
        self._tekma_combo.set("— izberi —")
        self._izbrana_tekma = None
        if hasattr(self, "_lbl_domaci"):
            self._lbl_domaci.configure(text="Goli domači")
            self._lbl_gostje.configure(text="Goli gostje")
        skupaj = len(vse_tekme)
        preostalo = len(self._tekme_data)
        self._log(self._tekma_log,
                  f"Naloženih {preostalo} neobdelanih tekem (od {skupaj} skupaj).")

    def _on_tekma_select(self, _=None):
        idx = self._tekma_combo.current()
        if idx < 0 or not hasattr(self, "_tekme_data"): return
        t = self._tekme_data[idx]
        self._izbrana_tekma = t
        self._lbl_domaci.configure(text=f"Goli domači  –  {t['ekipa1']}")
        self._lbl_gostje.configure(text=f"Goli gostje  –  {t['ekipa2']}")
        self._log(self._tekma_log,
                  f"Izbrana tekma: {t['ekipa1']} – {t['ekipa2']}  "
                  f"[stolpec: {t['stolpec']}, vrstica: {t.get('vrstica_napovedi', '–')}]")

    def _zacni_tekmo(self):
        try:
            gd  = int(self._t_domaci.get())
            gg  = int(self._t_gostje.get())
        except ValueError:
            messagebox.showerror("Napaka", "Preverite vnose (goli morajo biti številke).")
            return

        izbrana = getattr(self, "_izbrana_tekma", None)
        if not izbrana:
            messagebox.showerror("Napaka", "Najprej izberi tekmo iz seznama.")
            return
        vr = izbrana.get("vrstica_napovedi")
        if vr is None:
            messagebox.showerror("Napaka", "Vrstica napovedi ni določena v excelu (stolpec F).")
            return
        vr = int(vr)
        st = str(izbrana["stolpec"] or "").strip().upper()
        self._log(self._tekma_log, f"Stolpec rezultatov: {st}")
        self._log(self._tekma_log, f"\n▷  Začetek obdelave tekme …")

        self._tekma_progress["value"] = 0
        def _po_obdelavi(log_cb):
            def _progress(val):
                self._tekma_progress.configure(value=val)
                self.update_idletasks()
            obdelaj_tekmo(vr, gd, gg, st, log_cb, _progress)
            if izbrana and TEKME_FILE:
                try:
                    oznaci_tekmo_obdelano(izbrana["vrstica_excela"])
                    self._log(self._tekma_log,
                               f"✓  Tekma označena kot obdelana v excelu.")
                    self._osveži_tekme()
                    self._sprazni_tekmo()
                except Exception as e:
                    self._log(self._tekma_log,
                               f"⚠  Ni uspelo označiti tekme: {e}")

        self._run_threaded(_po_obdelavi,
                           lambda m: self._log(self._tekma_log, m))

    def _sprazni_tekmo(self):
        self._t_domaci.set("")
        self._t_gostje.set("")
        self._izbrana_tekma = None
        self._lbl_domaci.configure(text="Goli domači")
        self._lbl_gostje.configure(text="Goli gostje")

    # ── STRAN: SKUPINA ────────────────────────

    def _build_skupina(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Zaključi skupino",
                 bg=DARK, fg=FG, font=FONT_HEAD,
                 pady=16).grid(row=0, column=0, columnspan=2, sticky="w", padx=18)

        left = self._card(parent, "Parametri skupine", row=1, col=0)

        tk.Label(left, text="Skupina iz excela (opcijsko)", bg=CARD,
                 fg=FG2, font=FONT_LABEL).pack(anchor="w", pady=(6, 1))
        self._skupina_var = tk.StringVar(value="— izberi —")
        self._skupina_combo = ttk.Combobox(
            left, textvariable=self._skupina_var,
            state="readonly", font=FONT_BODY, width=48
        )
        self._skupina_combo.pack(fill="x", ipady=4)
        self._skupina_combo.bind("<<ComboboxSelected>>", self._on_skupina_select)
        self._btn(left, "↻  Osveži seznam", self._osveži_skupine).pack(
            anchor="w", pady=(4, 10))

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=8)

        tk.Label(left, text="Ekipe (1–4, vrstni red končnega stanja)",
                 bg=CARD, fg=FG2, font=FONT_LABEL).pack(anchor="w", pady=(10, 2))
        self._s_ekipe = []
        for i in range(4):
            v = tk.StringVar(value="— izberi —")
            cb = ttk.Combobox(left, textvariable=v, state="readonly",
                              font=FONT_BODY, width=28)
            cb.pack(fill="x", ipady=4, pady=1)
            cb.bind("<<ComboboxSelected>>", self._on_ekipa_select)
            self._s_ekipe.append((v, cb))
        self._s_ekipe_combos = [pair[1] for pair in self._s_ekipe]

        self._btn(left, "◈  Obdelaj skupino",
                  self._zacni_skupino, GREEN).pack(anchor="w", pady=(14, 0))

        right = self._card(parent, "Dnevnik", row=1, col=1)
        self._skupina_log = self._log_box(right, rows=16)
        self._skupina_progress = ttk.Progressbar(right, mode="determinate", maximum=100)
        self._skupina_progress.pack(fill="x", pady=(6, 0))

    def _osveži_skupine(self):
        vse_skupine = preberi_skupine_iz_excela()
        self._skupine_data = [s for s in vse_skupine if str(s["obdelana"] or "").upper() != "D"]
        vrednosti = [f"Skupina {s['skupina']}" for s in self._skupine_data]
        self._skupina_combo["values"] = vrednosti
        self._skupina_combo.set("— izberi —")
        self._izbrana_skupina = None
        skupaj = len(vse_skupine)
        preostalo = len(self._skupine_data)
        self._log(self._skupina_log,
                  f"Naloženih {preostalo} neobdelanih skupin (od {skupaj} skupaj).")

    def _on_skupina_select(self, _=None):
        idx = self._skupina_combo.current()
        if idx < 0 or not hasattr(self, "_skupine_data"): return
        s = self._skupine_data[idx]
        self._izbrana_skupina = s
        oznaka = str(s.get("skupina") or "").strip().upper()
        ekipe = SKUPINE_SP2026.get(oznaka, [])
        self._posodobi_ekipe_combo(ekipe)

    def _posodobi_ekipe_combo(self, ekipe):
        """Nastavi vrednosti vseh 4 dropdownov glede na ekipe izbrane skupine."""
        # Poiščemo zadnje 4 Comboboxe v levi kartici
        combos = self._s_ekipe_combos
        for cb in combos:
            cb["values"] = ekipe
            cb.set("— izberi —")
        for v, _ in self._s_ekipe:
            v.set("— izberi —")

    def _on_ekipa_select(self, _=None):
        """Ob vsaki spremembi dropdowna posodobi razpoložljive možnosti v ostalih."""
        izbrane = {v.get() for v, _ in self._s_ekipe if v.get() != "— izberi —"}
        sk = getattr(self, "_izbrana_skupina", None)
        if not sk: return
        oznaka = str(sk.get("skupina") or "").strip().upper()
        vse_ekipe = SKUPINE_SP2026.get(oznaka, [])
        combos = self._s_ekipe_combos
        for (var, cb) in self._s_ekipe:
            trenutna = var.get()
            mozne = ["— izberi —"] + [
                e for e in vse_ekipe
                if e not in izbrane or e == trenutna
            ]
            cb["values"] = mozne

    def _zacni_skupino(self):
        izbrana_s = getattr(self, "_izbrana_skupina", None)
        if not izbrana_s:
            messagebox.showerror("Napaka", "Najprej izberi skupino iz seznama.")
            return
        vr = izbrana_s.get("nap_vrstica")
        if vr is None:
            messagebox.showerror("Napaka", "Začetna vrstica ni določena v excelu (stolpec D).")
            return
        nst = str(izbrana_s.get("nap_stolpec") or "").strip().upper()
        if not nst:
            messagebox.showerror("Napaka", "Stolpec napovedi ni določen v excelu (stolpec C).")
            return
        rst = str(izbrana_s.get("stolpec") or "").strip().upper()
        if not rst:
            messagebox.showerror("Napaka", "Stolpec rezultatov ni določen v excelu (stolpec B).")
            return
        vr = int(vr)
        ekipe = [v.get().strip().upper() for v, _ in self._s_ekipe]
        if any(e in ("", "— IZBERI —") for e in ekipe):
            messagebox.showerror("Napaka", "Izberi vse 4 ekipe.")
            return
        if len(set(ekipe)) < 4:
            duplikati = [e for e in ekipe if ekipe.count(e) > 1]
            messagebox.showerror("Napaka", f"Vsaka ekipa sme biti izbrana samo enkrat.\nDuplikat: {duplikati[0].title()}")
            return
        sk  = str(izbrana_s.get("skupina") or "").strip().upper()
        self._log(self._skupina_log, f"\n◈  Začetek obdelave skupine {sk} …")

        self._skupina_progress["value"] = 0
        izbrana_s_ref = izbrana_s
        def _po_skupini(log_cb):
            def _progress(val):
                self._skupina_progress.configure(value=val)
                self.update_idletasks()
            obdelaj_skupino(sk, vr, nst, rst, ekipe, log_cb, _progress)
            if izbrana_s_ref.get("vrstica_excela") and TEKME_FILE:
                try:
                    oznaci_skupino_obdelano(izbrana_s_ref["vrstica_excela"])
                    self._log(self._skupina_log,
                               f"✓  Skupina {sk} označena kot obdelana v excelu.")
                    self._osveži_skupine()
                except Exception as e:
                    self._log(self._skupina_log,
                               f"⚠  Ni uspelo označiti skupine: {e}")
            self._sprazni_skupino()

        self._run_threaded(_po_skupini,
                           lambda m: self._log(self._skupina_log, m))

    def _sprazni_skupino(self):
        for v, _ in self._s_ekipe:
            v.set("— izberi —")
        self._izbrana_skupina = None
        self._skupina_combo.set("— izberi —")
        # Počisti vrednosti v dropdownih ekip
        combos = self._s_ekipe_combos
        for cb in combos:
            cb["values"] = []

    # ── STRAN: IMENA ──────────────────────────

    def _build_imena(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Naloži imena v excel",
                 bg=DARK, fg=FG, font=FONT_HEAD,
                 pady=16).grid(row=0, column=0, sticky="w", padx=18)

        card = self._card(parent, "Akcija", row=1, col=0)

        tk.Label(card, text=(
            "Prebere vsa imena iz map napovedi in jih zapiše\n"
            "v excel rezultatov (stolpec B, začenši z vrstico 3).\n\n"
            "Obstoječa vsebina v stolpcu B bo prepisana."
        ), bg=CARD, fg=FG2, font=FONT_LABEL, justify="left").pack(
            anchor="w", pady=(0, 12))

        self._btn(card, "⊕  Naloži imena",
                  self._zacni_imena, ACCENT).pack(anchor="w")

        self._imenа_log = self._log_box(card, rows=16)

    def _zacni_imena(self):
        # Safeguard: preveri, ali so podatki že vneseni
        def _preveri_in_naloži():
            ok, razlog = preveri_rezultati_prazni()
            if not ok:
                # Prikaži opozorilo in ponudi možnost nadaljevanja
                self._vprasaj_prepisovanje(razlog)
            else:
                self._log(self._imenа_log, "\n⊕  Nalaganje imen …")
                napolni_imena(lambda m: self._log(self._imenа_log, m))

        self._run_threaded(_preveri_in_naloži)

    def _vprasaj_prepisovanje(self, razlog):
        odg = messagebox.askyesno(
            "Pozor – podatki že obstajajo",
            f"{razlog}\n\nŽeliš vseeno prepisati stolpec B z imeni?\n"
            "(Točke v ostalih stolpcih ostanejo nespremenjene.)",
            icon="warning"
        )
        if odg:
            self._log(self._imenа_log, "\n⚠  Prepisovanje imen …")
            self._run_threaded(napolni_imena,
                               lambda m: self._log(self._imenа_log, m))
        else:
            self._log(self._imenа_log, "↩  Nalaganje preklicano.")

    # ── STRAN: SORTIRAJ ───────────────────────

    def _build_sortiraj(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Sortiraj in oštevilči lestvico",
                 bg=DARK, fg=FG, font=FONT_HEAD,
                 pady=16).grid(row=0, column=0, columnspan=2, sticky="w", padx=18)

        # Kartica – predtekmovanje
        left = self._card(parent, "Predtekmovanje", row=1, col=0)

        tk.Label(left, text=(
            "Razvrsti vse vrstice po skupnem seštevku točk (stolpec C)\n"
            "in doda ustrezne oznake mest v stolpec A.\n\n"
            "Pri enakem številu točk se mesto podvoji, vmesna mesta so prazna."
        ), bg=CARD, fg=FG2, font=FONT_LABEL, justify="left").pack(
            anchor="w", pady=(0, 12))

        self._btn(left, "⇅  Sortiraj predtekmovanje",
                  self._zacni_sortiranje, ACCENT).pack(anchor="w")

        self._sort_log = self._log_box(left, rows=14)

        # Kartica – izločilni del
        right = self._card(parent, "Izločilni del", row=1, col=1)

        tk.Label(right, text=(
            "Enako kot pri predtekmovanju, dodatno pa obarva\n"
            "imena igralcev rdeče, če nobena njihova napovedana\n"
            "ekipa ni več v igri (ni mogoče doseči novih točk)."
        ), bg=CARD, fg=FG2, font=FONT_LABEL, justify="left").pack(
            anchor="w", pady=(0, 12))

        self._btn(right, "⇅  Sortiraj izločilne",
                  self._zacni_sortiranje_izl, ACCENT).pack(anchor="w")

        self._sort_log_izl = self._log_box(right, rows=14)

    def _zacni_sortiranje(self):
        self._log(self._sort_log, "\n⇅  Sortiranje …")
        self._run_threaded(sortiraj_in_ostevilci,
                           log_cb=lambda m: self._log(self._sort_log, m))

    def _zacni_sortiranje_izl(self):
        self._log(self._sort_log_izl, "\n⇅  Sortiranje izločilnega dela …")
        self._run_threaded(
            sortiraj_in_ostevilci,
            rezultati_file=REZULTATI_FILE_IZL,
            napovedi_mapa=NAPOVEDI_MAPA_IZL,
            ime_cell=IME_CELL_IZL,
            preveri_izpadle=True,
            start_row=4,
            log_cb=lambda m: self._log(self._sort_log_izl, m)
        )

    # ── STRAN: PREVERI NAPOVEDI ──────────────────

    def _build_preveri(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(1, weight=1)
        parent.rowconfigure(2, weight=1)

        tk.Label(parent, text="Preveri napovedi",
                 bg=DARK, fg=FG, font=FONT_HEAD,
                 pady=16).grid(row=0, column=0, columnspan=2, sticky="w", padx=18)

        left = self._card(parent, "Preveri imena ekip v skupinah", row=1, col=0)

        tk.Label(left, text=(
            "Preveri, ali napovedi vseh uporabnikov\n"
            "vsebujejo prepoznavna imena ekip za vsako skupino.\n\n"
            "Referenčna imena so vgrajena v kodo (SP 2026)."
        ), bg=CARD, fg=FG2, font=FONT_LABEL, justify="left").pack(anchor="w", pady=(0, 16))

        tk.Label(left, text="Prag ujemanja (%)", bg=CARD, fg=FG2,
                 font=FONT_LABEL).pack(anchor="w", pady=(0, 1))
        self._preveri_prag = tk.StringVar(value=str(PRAG_VALIDACIJE))
        tk.Entry(left, textvariable=self._preveri_prag, bg=DARK, fg=FG,
                 insertbackground=FG, relief="flat", font=FONT_BODY, width=6).pack(anchor="w", ipady=4)

        tk.Label(left, text=(
            "Priporočeno: 75\n"
            "(nižje = manj strogo, višje = strožje)"
        ), bg=CARD, fg=MUTED, font=FONT_LABEL).pack(anchor="w", pady=(2, 16))

        self._btn(left, "🔍  Zaženi preverjanje",
                  self._zacni_validacijo, ACCENT).pack(anchor="w")

        right = self._card(parent, "Rezultati preverjanja", row=1, col=1)
        self._preveri_log = self._log_box(right, rows=14)

        # Drugi panel – popolnost vnosov izločilnih tekem
        left2 = self._card(parent, "Preveri popolnost vnosov – izločilni del", row=2, col=0)

        tk.Label(left2, text=(
            "Gre čez vse napovedi in preveri vse neobdelane tekme\n"
            "(tudi R8/R4/R2/FIN, kjer ekipi še nista znani) – preveri\n"
            "izpolnjenost in pravilnost imen glede na ekipe, ki teoretično\n"
            "lahko pridejo v ta del bracketa. Pri tekmi za 3. mesto se\n"
            "preverja samo izpolnjenost. Uporablja isti prag kot zgoraj."
        ), bg=CARD, fg=FG2, font=FONT_LABEL, justify="left").pack(anchor="w", pady=(0, 16))

        self._btn(left2, "🔍  Preveri izločilne vnose",
                  self._zacni_validacijo_izl, ACCENT).pack(anchor="w")

        right2 = self._card(parent, "Rezultati preverjanja – izločilni", row=2, col=1)
        self._preveri_izl_log = self._log_box(right2, rows=14)

    def _zacni_validacijo(self):
        global PRAG_VALIDACIJE
        try:
            PRAG_VALIDACIJE = int(self._preveri_prag.get())
        except ValueError:
            messagebox.showerror("Napaka", "Prag mora biti število med 0 in 100.")
            return

        # Poiščemo podatke skupin iz excela (stolpec napovedi, vrstica)
        skupine_excel = {str(s["skupina"]).strip().upper(): s
                         for s in preberi_skupine_iz_excela()}

        skupine_ekipe = []
        for oznaka, ekipe in SKUPINE_SP2026.items():
            sk_data = skupine_excel.get(oznaka)
            if sk_data is None:
                self._log(self._preveri_log,
                          f"⚠  Skupina {oznaka} ni najdena v excelu – preskočena.")
                continue
            skupine_ekipe.append({
                "skupina":     oznaka,
                "nap_stolpec": sk_data["nap_stolpec"],
                "nap_vrstica": sk_data["nap_vrstica"],
                "ekipe":       ekipe,
            })

        if not skupine_ekipe:
            messagebox.showerror("Napaka", "Nobena skupina ni bila najdena v excelu.")
            return

        self._log(self._preveri_log,
                  f"\n🔍  Začetek preverjanja {len(skupine_ekipe)} skupin …\n")
        self._run_threaded(validiraj_napovedi_skupin, skupine_ekipe,
                           lambda m: self._log(self._preveri_log, m))

    def _zacni_validacijo_izl(self):
        self._log(self._preveri_izl_log, "\n🔍  Začetek preverjanja izločilnih vnosov …\n")
        self._run_threaded(validiraj_napovedi_izlocilnih,
                           lambda m: self._log(self._preveri_izl_log, m))

    # ── STRAN: PRENESI V IZLOČILNE ────────────────

    def _build_prenesi(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Prenesi v izločilni del",
                 bg=DARK, fg=FG, font=FONT_HEAD,
                 pady=16).grid(row=0, column=0, sticky="w", padx=18)

        card = self._card(parent, "Akcija", row=1, col=0)

        tk.Label(card, text=(
            "Prenese imena in skupne točke vseh igralcev\n"
            "iz excela predtekmovanja v excel izločilnih bojev\n"
            "(stolpec A: mesto, B: ime, D: točke predtekmovanja).\n\n"
            "Obstoječa vsebina v izločilnem excelu bo prepisana."
        ), bg=CARD, fg=FG2, font=FONT_LABEL, justify="left").pack(
            anchor="w", pady=(0, 12))

        self._btn(card, "⤴  Prenesi podatke",
                  self._zacni_prenesi, ACCENT).pack(anchor="w")

        self._prenesi_log = self._log_box(card, rows=16)

    def _zacni_prenesi(self):
        self._log(self._prenesi_log, "\n⤴  Prenašanje podatkov …")
        self._run_threaded(prenesi_v_izlocilne,
                           lambda m: self._log(self._prenesi_log, m))

    # ── STRAN: VNESI IZLOČILNO TEKMO ──────────────

    def _build_izlocilna(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(1, weight=1)

        tk.Label(parent, text="Vnesi izločilno tekmo",
                 bg=DARK, fg=FG, font=FONT_HEAD,
                 pady=16).grid(row=0, column=0, columnspan=2, sticky="w", padx=18)

        left = self._card(parent, "Podatki tekme", row=1, col=0)

        tk.Label(left, text="Tekma iz excela", bg=CARD,
                 fg=FG2, font=FONT_LABEL).pack(anchor="w", pady=(6, 1))
        self._izl_var = tk.StringVar(value="— izberi —")
        self._izl_combo = ttk.Combobox(
            left, textvariable=self._izl_var,
            state="readonly", font=FONT_BODY, width=48
        )
        self._izl_combo.pack(fill="x", ipady=4)
        self._izl_combo.bind("<<ComboboxSelected>>", self._on_izl_select)
        self._btn(left, "↻  Osveži seznam", self._osveži_izlocilne).pack(
            anchor="w", pady=(4, 10))

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=8)

        self._izl_lbl_domaci = tk.Label(left, text="Goli ekipa 1", bg=CARD, fg=FG2, font=FONT_LABEL)
        self._izl_lbl_domaci.pack(anchor="w", pady=(6, 1))
        self._izl_gd = tk.StringVar()
        tk.Entry(left, textvariable=self._izl_gd, bg=DARK, fg=FG,
                 insertbackground=FG, relief="flat", font=FONT_BODY, width=22).pack(fill="x", ipady=5)

        self._izl_lbl_gostje = tk.Label(left, text="Goli ekipa 2", bg=CARD, fg=FG2, font=FONT_LABEL)
        self._izl_lbl_gostje.pack(anchor="w", pady=(6, 1))
        self._izl_gg = tk.StringVar()
        tk.Entry(left, textvariable=self._izl_gg, bg=DARK, fg=FG,
                 insertbackground=FG, relief="flat", font=FONT_BODY, width=22).pack(fill="x", ipady=5)

        self._izl_izenaceno = tk.BooleanVar(value=False)
        self._izl_check = tk.Checkbutton(
            left, text="Izenačeno – izberi kdo gre naprej",
            variable=self._izl_izenaceno, command=self._on_izl_izenaceno,
            bg=CARD, fg=FG2, selectcolor=DARK, activebackground=CARD,
            activeforeground=FG, font=FONT_LABEL
        )
        self._izl_check.pack(anchor="w", pady=(10, 2))

        self._izl_nap_var = tk.StringVar(value="— izberi —")
        self._izl_nap_combo = ttk.Combobox(
            left, textvariable=self._izl_nap_var,
            state="disabled", font=FONT_BODY, width=28
        )
        self._izl_nap_combo.pack(fill="x", ipady=4)

        self._btn(left, "🏆  Obdelaj tekmo",
                  self._zacni_izlocilno, GREEN).pack(anchor="w", pady=(14, 0))

        right = self._card(parent, "Dnevnik", row=1, col=1)
        self._izl_log = self._log_box(right, rows=16)
        self._izl_progress = ttk.Progressbar(right, mode="determinate", maximum=100)
        self._izl_progress.pack(fill="x", pady=(6, 0))

    def _osveži_izlocilne(self):
        vse = preberi_izlocilne_iz_excela()
        # Prikaži samo tiste z obema ekipama znanima in neobdelane
        self._izl_data = [
            t for t in vse
            if t["ekipa1"] and t["ekipa2"] and str(t["obdelana"] or "").upper() != "D"
        ]
        vrednosti = [
            f"{t['stage']} – {t['ekipa1']} : {t['ekipa2']}"
            for t in self._izl_data
        ]
        self._izl_combo["values"] = vrednosti
        self._izl_combo.set("— izberi —")
        self._izbrana_izl = None
        if hasattr(self, "_izl_lbl_domaci"):
            self._izl_lbl_domaci.configure(text="Goli ekipa 1")
            self._izl_lbl_gostje.configure(text="Goli ekipa 2")
        skupaj = len(vse)
        pripravljenih = len(self._izl_data)
        self._log(self._izl_log,
                  f"Naloženih {pripravljenih} pripravljenih tekem (od {skupaj} skupaj, "
                  f"ostale še čakajo na znane ekipe).")

    def _on_izl_select(self, _=None):
        idx = self._izl_combo.current()
        if idx < 0 or not hasattr(self, "_izl_data"): return
        t = self._izl_data[idx]
        self._izbrana_izl = t
        self._izl_lbl_domaci.configure(text=f"Goli ekipa 1  –  {t['ekipa1']}")
        self._izl_lbl_gostje.configure(text=f"Goli ekipa 2  –  {t['ekipa2']}")
        self._izl_nap_combo["values"] = [t["ekipa1"], t["ekipa2"]]
        self._izl_nap_combo.set("— izberi —")
        self._log(self._izl_log,
                  f"Izbrana tekma: {t['stage']} {t['ekipa1']} : {t['ekipa2']}  "
                  f"[stolpec: {t['stolpec_rez']}, točke napr.: {t['st_tock']}]")

    def _on_izl_izenaceno(self):
        if self._izl_izenaceno.get():
            self._izl_nap_combo.configure(state="readonly")
        else:
            self._izl_nap_combo.configure(state="disabled")
            self._izl_nap_combo.set("— izberi —")

    def _zacni_izlocilno(self):
        try:
            gd = int(self._izl_gd.get())
            gg = int(self._izl_gg.get())
        except ValueError:
            messagebox.showerror("Napaka", "Preverite vnose (goli morajo biti številke).")
            return

        izbrana = getattr(self, "_izbrana_izl", None)
        if not izbrana:
            messagebox.showerror("Napaka", "Najprej izberi tekmo iz seznama.")
            return

        if gd == gg:
            if not self._izl_izenaceno.get():
                messagebox.showerror("Napaka", "Rezultat je izenačen – obkljukaj 'Izenačeno' in izberi kdo gre naprej.")
                return
            napredujoca = self._izl_nap_var.get()
            if napredujoca not in (izbrana["ekipa1"], izbrana["ekipa2"]):
                messagebox.showerror("Napaka", "Izberi katera ekipa gre naprej.")
                return
        else:
            napredujoca = izbrana["ekipa1"] if gd > gg else izbrana["ekipa2"]

        self._log(self._izl_log, f"\n🏆  Začetek obdelave tekme {izbrana['stage']} …")
        self._izl_progress["value"] = 0

        def _po_obdelavi(log_cb):
            def _progress(val):
                self._izl_progress.configure(value=val)
                self.update_idletasks()
            obdelaj_izlocilno_tekmo(izbrana, gd, gg, napredujoca, log_cb, _progress)
            try:
                zapisi_napredujočo_ekipo(izbrana["vrstica_excela"], napredujoca, log_cb)
                oznaci_izlocilno_obdelano(izbrana["vrstica_excela"])
                self._log(self._izl_log,
                           f"✓  Tekma označena kot obdelana, {napredujoca} napreduje.")
                self._osveži_izlocilne()
                self._sprazni_izlocilno()
            except Exception as e:
                self._log(self._izl_log,
                           f"⚠  Napaka pri zaključevanju: {e}")

        self._run_threaded(_po_obdelavi,
                           lambda m: self._log(self._izl_log, m))

    def _sprazni_izlocilno(self):
        self._izl_gd.set("")
        self._izl_gg.set("")
        self._izbrana_izl = None
        self._izl_izenaceno.set(False)
        self._izl_nap_combo.configure(state="disabled")
        self._izl_nap_combo.set("— izberi —")
        self._izl_lbl_domaci.configure(text="Goli ekipa 1")
        self._izl_lbl_gostje.configure(text="Goli ekipa 2")

    # ── STRAN: NASTAVITVE ─────────────────────

    def _build_nastavitve(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(4, weight=1)

        tk.Label(parent, text="Nastavitve",
                 bg=DARK, fg=FG, font=FONT_HEAD,
                 pady=16).grid(row=0, column=0, columnspan=2, sticky="w", padx=18)

        # Kartica – poti do datotek
        paths = self._card(parent, "Poti do datotek", row=1, col=0)
        self._n_napovedi   = self._labeled_entry(paths, "Mapa napovedi",       NAPOVEDI_MAPA, 36)
        self._n_rezultati  = self._labeled_entry(paths, "Excel rezultatov",    REZULTATI_FILE, 36)
        self._n_tekme_file = self._labeled_entry(paths, "Excel tekem/skupin",  TEKME_FILE, 36)

        # Kartica – stolpci napovedi
        cols = self._card(parent, "Stolpci v napovedi", row=1, col=1)
        self._n_ime_cell   = self._labeled_entry(cols, "Celica z imenom",   IME_CELL,   12)
        self._n_domaci_col = self._labeled_entry(cols, "Stolpec domači",    DOMACI_COL, 12)
        self._n_gost_col   = self._labeled_entry(cols, "Stolpec gostje",    GOST_COL,   12)
        self._n_tocke_col  = self._labeled_entry(cols, "Stolpec točke",     TOCKE_COL,  12)

        # Kartica – stolpci rezultatov
        rcols = self._card(parent, "Stolpci v rezultatih", row=2, col=0)
        self._n_ime_col    = self._labeled_entry(rcols, "Stolpec imen",     IME_COLUMN, 12)
        self._n_tek_sheet  = self._labeled_entry(rcols, "Indeks lista – tekme",   "0", 6)
        self._n_sk_sheet   = self._labeled_entry(rcols, "Indeks lista – skupine", "1", 6)
        self._n_izl_sheet  = self._labeled_entry(rcols, "Indeks lista – izločilni", "2", 6)

        # Kartica – izločilni del
        izl = self._card(parent, "Izločilni del", row=2, col=1)
        self._n_napovedi_izl  = self._labeled_entry(izl, "Mapa napovedi (izločilni)",    NAPOVEDI_MAPA_IZL, 36)
        self._n_rezultati_izl = self._labeled_entry(izl, "Excel rezultatov (izločilni)", REZULTATI_FILE_IZL, 36)
        self._n_ime_cell_izl  = self._labeled_entry(izl, "Celica z imenom (izločilni)",  IME_CELL_IZL, 12)
        self._n_ime_col_izl   = self._labeled_entry(izl, "Stolpec imen (izločilni)",     IME_COLUMN_IZL, 12)

        # Gumbi
        btn_row = tk.Frame(parent, bg=DARK)
        btn_row.grid(row=3, column=1, sticky="sw", padx=8, pady=6)
        self._btn(btn_row, "✓  Shrani nastavitve",
                  self._shrani_nastavitve, GREEN).pack(side="left", padx=(0, 8))
        self._btn(btn_row, "↺  Ponastavi",
                  self._refresh_globals, RED).pack(side="left")

        # Status
        self._n_status = tk.Label(parent, text="", bg=DARK, fg=GREEN,
                                  font=FONT_LABEL)
        self._n_status.grid(row=4, column=0, columnspan=2,
                            sticky="w", padx=18, pady=4)

    def _shrani_nastavitve(self):
        global NAPOVEDI_MAPA, REZULTATI_FILE, TEKME_FILE
        global IME_CELL, DOMACI_COL, GOST_COL, TOCKE_COL, IME_COLUMN
        global TEKME_SHEET_IDX, SKUPINE_SHEET_IDX, IZLOCILNI_SHEET_IDX
        global NAPOVEDI_MAPA_IZL, REZULTATI_FILE_IZL, IME_CELL_IZL, IME_COLUMN_IZL

        NAPOVEDI_MAPA      = self._n_napovedi.get().strip()
        REZULTATI_FILE     = self._n_rezultati.get().strip()
        TEKME_FILE         = self._n_tekme_file.get().strip()
        IME_CELL           = self._n_ime_cell.get().strip().upper()
        DOMACI_COL         = self._n_domaci_col.get().strip().upper()
        GOST_COL           = self._n_gost_col.get().strip().upper()
        TOCKE_COL          = self._n_tocke_col.get().strip().upper()
        IME_COLUMN         = self._n_ime_col.get().strip().upper()
        NAPOVEDI_MAPA_IZL  = self._n_napovedi_izl.get().strip()
        REZULTATI_FILE_IZL = self._n_rezultati_izl.get().strip()
        IME_CELL_IZL       = self._n_ime_cell_izl.get().strip().upper()
        IME_COLUMN_IZL     = self._n_ime_col_izl.get().strip().upper()
        try:
            TEKME_SHEET_IDX     = int(self._n_tek_sheet.get())
            SKUPINE_SHEET_IDX   = int(self._n_sk_sheet.get())
            IZLOCILNI_SHEET_IDX = int(self._n_izl_sheet.get())
        except ValueError:
            pass

        # Osveži slovarja imen v ozadju
        self._n_status.configure(text="Nastavitve shranjene. Osveževanje slovarjev imen …", fg=FG2)
        self.update_idletasks()
        try:
            napolni_slovar_imen()
            napolni_slovar_imen_izl()
            self._n_status.configure(
                text=f"✓  Naloženih {len(IMENA_VRSTICE)} imen (predtekm.), "
                     f"{len(IMENA_VRSTICE_IZL)} imen (izločilni).", fg=GREEN)
        except Exception as e:
            self._n_status.configure(
                text=f"✗  Napaka pri nalaganju: {e}", fg=RED)
        shrani_config()

    def _refresh_globals(self):
        """Nastavi Entry vrednosti iz trenutnih globalnih spremenljivk."""
        pairs = [
            (self._n_napovedi,   NAPOVEDI_MAPA),
            (self._n_rezultati,  REZULTATI_FILE),
            (self._n_tekme_file, TEKME_FILE),
            (self._n_ime_cell,   IME_CELL),
            (self._n_domaci_col, DOMACI_COL),
            (self._n_gost_col,   GOST_COL),
            (self._n_tocke_col,  TOCKE_COL),
            (self._n_ime_col,    IME_COLUMN),
            (self._n_tek_sheet,  str(TEKME_SHEET_IDX)),
            (self._n_sk_sheet,   str(SKUPINE_SHEET_IDX)),
            (self._n_izl_sheet,  str(IZLOCILNI_SHEET_IDX)),
            (self._n_napovedi_izl,  NAPOVEDI_MAPA_IZL),
            (self._n_rezultati_izl, REZULTATI_FILE_IZL),
            (self._n_ime_cell_izl,  IME_CELL_IZL),
            (self._n_ime_col_izl,   IME_COLUMN_IZL),
        ]
        # Entry varibale so dostopne le po inicializaciji
        try:
            for var, val in pairs:
                var.set(val)
        except AttributeError:
            pass  # še ni zgrajeno


# ──────────────────────────────────────────────
#  VSTOPNA TOČKA
# ──────────────────────────────────────────────

if __name__ == "__main__":
    # Nastavimo style za ttk Combobox
    style = None
    app   = SP2026App()

    s = ttk.Style(app)
    s.theme_use("clam")
    s.configure("TCombobox",
                 fieldbackground=DARK,
                 background=CARD,
                 foreground=FG,
                 selectbackground=DARK,
                 selectforeground="#ffffff",
                 insertcolor=FG,
                 arrowcolor=FG2)
    s.map("TCombobox",
          fieldbackground=[("readonly", DARK), ("disabled", PANEL)],
          foreground=[("readonly", FG), ("disabled", FG2)],
          selectbackground=[("readonly", DARK)],
          selectforeground=[("readonly", "#ffffff")])
	  
    napolni_slovar_imen()

    app.mainloop()
