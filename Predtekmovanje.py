import os
import openpyxl
import xlwings as xw
import logging
from rapidfuzz import process
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
import numpy as np

logging.basicConfig(
    filename="log.txt",
    level=logging.INFO,
    format="%(asctime)s - %(message)s"
)

NAPOVEDI_MAPA = "TESTNapovediPredtekmovanje"
REZULTATI_FILE = "TEST - SP 2026 - predtekmovanje - rezultati.xlsx"
LOG_FILE = "log.txt"

#Napovedi excel
IME_CELL = "I3"
DOMACI_COL = "I"
GOST_COL = "K"
TOCKE_COL = "L"

#Rezulati excel
IME_COLUMN = "B"

#Slovar imen in vrstic v katerih se nahajajo, da ne rabim vsakega imena iskat posebaj
IMENA_VRSTICE = {}

skupine = {
    "A": 6,
    "B": 13,
    "C": 20,
    "D": 27,
    "E": 34, 
    "F": 41,
    "G": 48,
    "H": 55,
    "I": 62,
    "J": 69,
    "K": 76,
    "L": 83
}


def vrni_tip(goliDomaci, goliGostje):

    if goliDomaci > goliGostje:
        return 1

    elif goliDomaci < goliGostje:
        return 2

    return 0


def napolni_slovar_imen():
    vrstica = 3

    appRezultati = xw.App(visible=False)
    wbRezultati = appRezultati.books.open(REZULTATI_FILE)
    wsRezultati = wbRezultati.sheets[0]

    global IMENA_VRSTICE
    imena = wsRezultati.range(f"{IME_COLUMN}{vrstica}").expand('down').value

    if imena is not None:
        IMENA_VRSTICE = {ime.strip().title(): i for i, ime in enumerate(imena, start=3)}  
    else:
        print("V rezultatih še ni imen!")

    wbRezultati.close()
    appRezultati.quit()


def poisci_skupino(ws, skupina):

    target = f"SKUPINA {skupina}"

    for row in range(1, ws.max_row + 1):

        value = ws[f"A{row}"].value

        if value == target:
            return row

    return None


def obdelaj_tekmo():

    print("\n=== VNOS REZULTATA TEKME ===")
 
    #naberem potrebne podatke
    vrsticaNapovedi = input("Vrstica: ").strip()

    goliDomaciRezultat = int(input(f"Goli domaci: "))
    goliGostjeRezultat = int(input(f"Goli gostje: "))

    stolpecRezultati = input("Stolpec v datoteki rezultatov: ")

    tipRezultat = vrni_tip(goliDomaciRezultat, goliGostjeRezultat)

    files = [
        f for f in os.listdir(NAPOVEDI_MAPA)
        if f.endswith(".xlsx") and not f.startswith("~")
    ]

    try:
        appRezultati = xw.App(visible=False)
        wbRezultati = appRezultati.books.open(REZULTATI_FILE)
        wsRezultati = wbRezultati.sheets[0]

        for file in files:
  
            path = os.path.join(NAPOVEDI_MAPA, file)

            try:
                wbNapoved = openpyxl.load_workbook(path, read_only=True, data_only=True)
                wsNapoved = wbNapoved.worksheets[0]

                ime = wsNapoved[IME_CELL].value.strip().title()
                vrsticaRezultati = IMENA_VRSTICE.get(ime)
                if vrsticaRezultati is None:
                    raise Exception("Ime ni najdeno!")

                goliDomaciNapoved = int(wsNapoved[f"{DOMACI_COL}{vrsticaNapovedi}"].value)
                goliGostjeNapoved = int(wsNapoved[f"{GOST_COL}{vrsticaNapovedi}"].value)

                tipNapoved = vrni_tip(goliDomaciNapoved, goliGostjeNapoved)

                if goliDomaciNapoved is None or goliDomaciNapoved is None:
                    raise Exception("Napaka branja napovedi - ni enega izmed števila golov!")

                #izracunam tocke
                if (goliDomaciNapoved == goliDomaciRezultat) and (goliGostjeNapoved == goliGostjeRezultat):
                    tocke = 3
                elif (tipNapoved == tipRezultat):
                    tocke = 1
                else:
                    tocke = 0

                wsRezultati[f"{stolpecRezultati}{vrsticaRezultati}"].value = tocke

                wbNapoved.close()

                logging.info(f"{ime} | {goliDomaciNapoved}:{goliGostjeRezultat} | {tocke} tock")

                print(f"OK -> {ime}: {tocke}")

            except Exception as e:
                logging.error(f"NAPAKA {file}: {e}")
                print(f"Napaka pri: {file}")
    finally:
        wbRezultati.save(REZULTATI_FILE)
        wbRezultati.close()
        appRezultati.quit()


def obdelaj_skupino():

    print("\n=== SKUPINSKI DEL ===")

    skupina = input("Skupina (A-L): ").upper().strip()
    vrsticaSkupineNapoved = int(input("Vrstica skupine v napovedi: "))
    stolpecSkupineNapoved = input("Stolpec skupine v napovedi: ")

    stolpecSkupineRezultati = input("Prvi stolpec skupine v rezultatih: ")

    drzava1 = input("Prva ekipa: ")
    drzava2 = input("Druga ekipa: ")
    drzava3 = input("Tretja ekipa: ")
    drzava4 = input("Četrta ekipa: ")

    ekipeRezultat = [drzava1, drzava2, drzava3, drzava4]

    ekipeRezultat = [x.strip().upper() for x in ekipeRezultat]

    if len(ekipeRezultat) != 4:
        print("Napaka: vnesti moraš 4 ekipe")
        return
    
    ekipeSeznam = np.array(ekipeRezultat).tolist()
    print(ekipeSeznam)

    files = [
        f for f in os.listdir(NAPOVEDI_MAPA)
        if f.endswith(".xlsx") and not f.startswith("~")
    ]

    appRezultati = xw.App(visible=False)
    wbRezultati = appRezultati.books.open(REZULTATI_FILE)
    wsRezultati = wbRezultati.sheets[0]

    try:
        for file in files:

            path = os.path.join(NAPOVEDI_MAPA, file)
            try:
                wbNapovedi = openpyxl.load_workbook(path, read_only=True, data_only=True)
                wsNapovedi = wbNapovedi.worksheets[0]

                ime = wsNapovedi[IME_CELL].value.strip().title()
                vrsticaRezultati = IMENA_VRSTICE.get(ime)
                if vrsticaRezultati is None:
                    raise Exception("Ime ni najdeno!")

                points = 0

                for idx in range(4):

                    napoved = wsNapovedi[f"{stolpecSkupineNapoved}{vrsticaSkupineNapoved + idx}"].value
                    napoved = napoved.upper()

                    match, procent, _ = process.extractOne(
                                        napoved,
                                        ekipeSeznam
                                    )
                    
                    if match == ekipeRezultat[idx]:
                        points += 2

                        #zapisem rezultat
                        indeksStolpec = column_index_from_string(stolpecSkupineRezultati)
                        indeksStolpec = get_column_letter(indeksStolpec + idx)

                        wsRezultati[f"{indeksStolpec}{vrsticaRezultati}"].value = 2


                logging.info(f"{ime} | Skupina {skupina} | {points} točk")

                print(f"OK -> {ime}: {points}")

            except Exception as e:

                logging.error(f"NAPAKA {file}: {e}")
                print(f"Napaka pri: {file}")

            wbNapovedi.close()
            
    finally:
        print("Konec")
        wbRezultati.save(REZULTATI_FILE)
        wbRezultati.close()
        appRezultati.quit()    

def napolni_imena_rezultati():
    #zacnem v 3 vrstici
    vrstica = 3

    files = [
        f for f in os.listdir(NAPOVEDI_MAPA)
        if f.endswith(".xlsx") and not f.startswith("~")
    ]

    appRezultati = xw.App(visible=False)
    wbRezultati = appRezultati.books.open(REZULTATI_FILE)
    wsRezultati = wbRezultati.sheets[0]

    for file in files:

        path = os.path.join(NAPOVEDI_MAPA, file)
        wbNapoved = None
        try:
            wbNapoved = openpyxl.load_workbook(path, read_only=True, data_only=True)
            wsNapoved = wbNapoved.worksheets[0]

            ime = wsNapoved[IME_CELL].value.strip().title()
            wbNapoved.close()

            wsRezultati[IME_COLUMN + str(vrstica)].value = ime

            vrstica = vrstica+1
            print(f"{file}: OK")

        except Exception as e:
            if wbNapoved is not None: 
                wbNapoved.close()
            logging.error(f"NAPAKA {file}: {e}")
            print(f"Napaka pri: {file}")
        
    wbRezultati.save(REZULTATI_FILE)
    wbRezultati.close()
    appRezultati.quit()
 
    napolni_slovar_imen()


def sortiraj_in_ostevilci():
    try:
        appRezultati = xw.App(visible=False)
        wbRezultati = appRezultati.books.open(REZULTATI_FILE)
        wsRezultati = wbRezultati.sheets[0]

        # Poiščemo zadnjo vrstico
        last_row = wsRezultati.used_range.last_cell.row

        # Poiščemo zadnji stolpec
        last_col = wsRezultati.used_range.last_cell.column

        # Preberemo vse igralce (brez glave)
        igralci = []

        try:
            for row in range(3, last_row + 1):

                vrstica = wsRezultati.range(
                    (row, 1),
                    (row, last_col)
                ).value

                # Preskoči prazne vrstice
                if vrstica[1] is None:
                    continue

                igralci.append(vrstica)

            # Sortiranje po stolpcu C (Skupaj)
            igralci.sort(
                key=lambda x: (
                    x[2] if isinstance(x[2], (int, float)) else 0
                ),
                reverse=True
            )

            # Zapišemo nazaj
            for idx, vrstica in enumerate(igralci, start=3):

                wsRezultati.range(
                    (idx, 1),
                    (idx, last_col)
                ).value = vrstica
        except Exception as e:
            logging.error(f"Napaka pri sortiranju: {e}")
            print(f"Napaka pri sortiranju!")
            exit()    

        # --------------------------------
        # Izračun mest
        # --------------------------------

        trenutno_mesto = 1
        prejsnje_tocke = None

        try:
            for row in range(3, len(igralci) + 2):

                tocke = wsRezultati.range((row, 3)).value

                if row == 3:
                    wsRezultati.range((row, 1)).value = str(trenutno_mesto)+"."
                else:
                    if tocke == prejsnje_tocke:
                        # enako število točk
                        wsRezultati.range((row, 1)).value = ""
                    else:
                        trenutno_mesto = row - 1
                        wsRezultati.range((row, 1)).value = str(trenutno_mesto)+"."

                prejsnje_tocke = tocke

        except Exception as e:
            logging.error(f"Napaka pri oštevilčevanju: {e}")
            print(f"Napaka pri oštevilčevanju!") 

    finally:
        wbRezultati.save(REZULTATI_FILE)
        wbRezultati.close()
        appRezultati.quit()

    print(f"Uspešno razvrščena in oštevilčena lestvica!") 

    IMENA_VRSTICE.clear
    napolni_slovar_imen()


def main():
    #pripravi slovar imen in vrstic v katerih so
    napolni_slovar_imen()

    while True:

        print("\n============================")
        print("SP 2026 - TOČKOVANJE")
        print("============================")
        print("1 - Vnesi rezultat tekme")
        print("2 - Zaključi skupino")
        print("3 - Napolni excel z imeni")
        print("4 - Sortiraj in oštevilči lestvico")
        print("X - Izhod")

        choice = input("Izbira: ").strip()

        if choice == "1":
            obdelaj_tekmo()

        elif choice == "2":
            obdelaj_skupino()        
            
        elif choice == "3":
            napolni_imena_rezultati()

        elif choice == "4":
            sortiraj_in_ostevilci()

        elif choice == "X":
            break

        else:
            print("Napačna izbira")


if __name__ == "__main__":
    main()